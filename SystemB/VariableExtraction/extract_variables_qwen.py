#!/usr/bin/env python3
"""
extract_variables_qwen.py  —  System B Variable Extraction (Qwen Q&A version)
Location: ~/digitize_medical_records/VariableExtraction/extract_variables_qwen.py

Replaces regex patterns with Qwen 72B/30B structured Q&A extraction.
Any scientist can add new variables by editing extraction_fields.yaml —
no regex knowledge needed, just write questions in plain language.

Usage:
  python extract_variables_qwen.py --patient patient_002_anonymized.txt
  python extract_variables_qwen.py --all
  python extract_variables_qwen.py --all --model qwen3:30b --host http://127.0.0.1:11435

100% local — no data leaves your machine.
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml
import ollama

# ── Paths ──────────────────────────────────────────────────────────────────────
WORK_DIR    = Path("/home/lfarooq/digitize_medical_records")
FINAL_ANON  = WORK_DIR / "outputs" / "final_anonymized"
OUTPUT_DIR  = WORK_DIR / "VariableExtraction" / "outputs"
FIELDS_PATH = WORK_DIR / "VariableExtraction" / "extraction_fields.yaml"

# ── Ollama config ──────────────────────────────────────────────────────────────
OLLAMA_HOST  = "http://127.0.0.1:11435"  # personal GPU server
DEFAULT_MODEL = "qwen3:30b"

# ── Column list (same 130 variables as before) ─────────────────────────────────
COLUMNS = [
    "ID_Patient","Sexe","Age","Pays_de_naissance","Durée_zone_endémie",
    "Date_départ_séjour","Date_retour_séjour","Lieu_séjour","Symptômes_séjour",
    "type_symptômes_séjour","Traitement_séjour","Drépanocytose","Splénectomie",
    "palu_ant","diabète","insuff_rénale","chimioprophylaxie","moustiquaire_imp",
    "consultation_pré_voyage","date_premiers_symptomes","méd_traitant",
    "Date_PEC_J0","fièvre_J0","anorexie_J0","diarrhées_J0","vomissement_J0",
    "nausée_J0","douleurs_abdominales_J0","asthénie_J0","trouble_conscience_J0",
    "convulsion_J0","céphalées_J0","autres_motifs_J0","Poids_J0","Température_J0",
    "Fréquence_Cardiaque_J0","SaO2_J0","Hépatomégalie_J0","spénomégalie_J0",
    "pâleur_J0","ictère_J0","raideur_nuque_J0","Goutte_épaisse_J0",
    "Frottis_sanguin_J0","Parasitémie_J0","PCR_J0","HRP2_J0","hémoculture_J0",
    "hémoglobine_J0","plaquettes_J0","globules_blancs_J0","eosinophiles_J0",
    "ASAT_J0","ALAT_J0","bilirubine_J0","Urée_J0","Créatinine_J0","CRP_J0",
    "PCT_J0","pH_J0","lactates_J0","TP_J0","TCA_J0",
    "Date_PEC_J3","fièvre_J3","anorexie_J3","diarrhées_J3","vomissement_J3",
    "nausée_J3","douleurs_abdominales_J3","asthénie_J3","trouble_conscience_J3",
    "convulsion_J3","céphalées_J3","autres_motifs_J3","Poids_J3","Température_J3",
    "Fréquence_Cardiaque_J3","SaO2_J3","Hépatomégalie_J3","spénomégalie_J3",
    "pâleur_J3","ictère_J3","raideur_nuque_J3","Goutte_épaisse_J3",
    "Frottis_sanguin_J3","Parasitémie_J3","PCR_J3","HRP2_J3","hémoculture_J3",
    "hémoglobine_J3","plaquettes_J3","globules_blancs_J3","eosinophiles_J3",
    "ASAT_J3","ALAT_J3","bilirubine_J3","Urée_J3","Créatinine_J3","CRP_J3",
    "PCT_J3","pH_J3","lactates_J3","TP_J3","TCA_J3",
    "PDV_J30","Date_PEC_J30","fièvre_J30","anorexie_J30","diarrhées_J30",
    "vomissement_J30","nausée_J30","douleurs_abdominales_J30","asthénie_J30",
    "trouble_conscience_J30","convulsion_J30","céphalées_J30","autres_motifs_J30",
    "Poids_J30","Température_J30","Fréquence_Cardiaque_J30","SaO2_J30",
    "Hépatomégalie_J30","spénomégalie_J30","pâleur_J30","ictère_J30",
    "raideur_nuque_J30","Goutte_épaisse_J30","Frottis_sanguin_J30",
    "Parasitémie_J30","PCR_J30","HRP2_J30","hémoculture_J30","hémoglobine_J30",
    "plaquettes_J30","globules_blancs_J30","eosinophiles_J30","ASAT_J30",
    "ALAT_J30","bilirubine_J30","Urée_J30","Créatinine_J30","CRP_J30",
    "PCT_J30","pH_J30","lactates_J30","TP_J30","TCA_J30",
    "Type_paludisme","gravité_palu","Traitement_antipalu","Posologie_traitement",
    "Durée_traitement","Perfusion","Hospitalisation","Durée_hospit",
    "type_service","ECG",
]


# ─────────────────────────────────────────────────────────────────────────────
# Utilities
# ─────────────────────────────────────────────────────────────────────────────
def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def parse_date(s: str) -> Optional[datetime]:
    s = str(s).strip().replace(".", "/")
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def fmt_date(s: str) -> str:
    d = parse_date(str(s).strip())
    return d.strftime("%d/%m/%Y") if d else str(s).strip()


def load_fields(path: Path) -> dict:
    if not path.exists():
        print(f"⚠️  Fields file not found: {path}")
        return {}
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — Parse anonymized file into sections
# ─────────────────────────────────────────────────────────────────────────────
def parse_anonymized_file(path: Path) -> dict:
    """
    Parse patient_NNN_anonymized.txt into:
      - sections: list of {doc_id, doc_type, cnr_fields, pages}
      - full_text: all text concatenated
    """
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()

    sections = []
    current = None
    in_cnr = False
    page_num = None
    page_lines = []

    doc_re  = re.compile(r"^DOCUMENT:\s+(\S+)\s+\[(\w+)\]$")
    page_re = re.compile(r"^===== page-(\d+) =====$")

    def flush_page():
        if current is not None and page_num is not None:
            current["pages"].append({
                "num": page_num,
                "text": "\n".join(page_lines)
            })

    for line in lines:
        m = doc_re.match(line.strip())
        if m:
            flush_page()
            page_lines.clear()
            page_num = None
            in_cnr = False
            current = {
                "doc_id": m.group(1),
                "doc_type": m.group(2),
                "cnr_fields": {},
                "pages": []
            }
            sections.append(current)
            continue

        if current is None:
            continue

        if "── CNR FORM FIELDS" in line:
            flush_page()
            page_lines.clear()
            page_num = None
            in_cnr = True
            continue

        if "── CLINICAL / LAB DOCUMENTS" in line:
            in_cnr = False
            continue

        if in_cnr:
            if ":" in line:
                k, _, v = line.partition(":")
                k, v = k.strip(), v.strip()
                if k and v and v.lower() not in ("none", ""):
                    current["cnr_fields"][k] = v
            continue

        mp = page_re.match(line.strip())
        if mp:
            flush_page()
            page_lines.clear()
            page_num = int(mp.group(1))
            continue

        if page_num is not None:
            page_lines.append(line)

    flush_page()

    return {
        "sections": sections,
        "full_text": text,
        "all_cnr": [s for s in sections if s["cnr_fields"]],
        "clinical_text": "\n".join(
            "\n".join(p["text"] for p in s["pages"])
            for s in sections
            if s["doc_type"] in ("NON_CNR", "MIXED")
        )
    }


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — Assign timepoints J0/J3/J30
# ─────────────────────────────────────────────────────────────────────────────
_DATE_PATS = [
    r"Date\s*:\s*(\d{1,2}/\d{2}/\d{2,4})",
    r"Date\s*/\s*Heure\s+(\d{2}/\d{2}/\d{4})",
    r"Pr.l.vement\s+pr.vu\s+le[^:]*:\s*(\d{2}/\d{2}/\d{2,4})",
    r"Pr.l.vement\s*:\s*(\d{2}/\d{2}/\d{4})",
    r"Entr.e Urgence le\s*(\d{2}/\d{2}/\d{4})",
    r"PARAMETRES A L.ARRIVEE\s*:\s*(\d{2}/\d{2}/\d{4})",
    r"COMPTE RENDU DE CONSULTATION DU\s+(\d{2}/\d{2}/\d{2,4})",
    r"COMPTE RENDU DE SEJOUR.*?DU\s+(\d{2}/\d{2}/\d{2,4})\s+AU",
    r"Date\s*:\s*(\d{2}\.\d{2}\.\d{4})",
    r"Date\s+Heure\s+(\d{2}/\d{2}/\d{4})",
]
_IGNORE_CTX = ["naissance", "naiss", "depart", "retour", "premiers", "derniere", "premiere"]


def page_date(text: str) -> str:
    for pat in _DATE_PATS:
        for m in re.finditer(pat, text, re.IGNORECASE):
            ctx = norm(text[max(0, m.start()-60):m.start()])
            if any(ig in ctx for ig in _IGNORE_CTX):
                continue
            d = parse_date(m.group(1))
            if d and 2000 <= d.year <= 2035:
                return d.strftime("%d/%m/%Y")
    return ""


def _has_explicit_j3_label(pages: list) -> bool:
    """Check if any page is explicitly labeled as J3/J4 follow-up."""
    j3_patterns = [
        r"cont[r]?[oô]le\s*j\s*[34]",   # handles OCR typo Contole J3
        r"suivi\s*j\s*[34]",
        r"j\s*[34]\s*(?:de|du|pour|palu|paludisme|cont[r]?[oô]le)",
        r"\bj\s*[34]\s*(?:frottis|parasit|nfs|bilan)",
        r"revient\s*[\xe0a]\s*j\s*[34]",
        r"retour\s*[\xe0a]\s*j\s*[34]",
        r"controle\s*j\s*[34]",
    ]
    for _, pg in pages:
        text_lower = pg["text"].lower()
        for pat in j3_patterns:
            if re.search(pat, text_lower):
                return True
    return False


def _lab_density(pages: list) -> int:
    """Count lab values in pages for data density scoring."""
    text = " ".join(pg["text"] for _, pg in pages)
    return len(re.findall(
        r'\b\d+[,.]?\d*\s*(?:g/dl|g/100|10x\d|mmol|ui/l|mg/l|%)\b',
        text, re.IGNORECASE))


def assign_timepoints(parsed: dict, j0_merge_days: int = 3) -> dict:
    """Group pages by date and assign J0/J3/J30.

    J3 selection rules (scientist, May 2026):
    1. EXPLICIT LABEL: page labeled 'Controle J3', 'Suivi J3' etc → that date is J3
    2. DATA DENSITY: pick date with most lab results in J0+2 to J0+5 window
    """
    date_pages: dict[str, list] = {}
    undated = []
    for sec in parsed["sections"]:
        for pg in sec["pages"]:
            d = page_date(pg["text"])
            if d:
                date_pages.setdefault(d, []).append((sec, pg))
            else:
                undated.append((sec, pg))
    sorted_dates = sorted(date_pages.keys(),
                          key=lambda d: parse_date(d) or datetime.min)
    groups = {"J0": [], "J3": [], "J30": []}
    dates  = {"J0": "", "J3": "", "J30": ""}
    n = len(sorted_dates)
    if n == 0:
        for sec in parsed["sections"]:
            for pg in sec["pages"]:
                groups["J0"].append(pg["text"])
    elif n == 1:
        groups["J0"] = [pg["text"] for _, pg in date_pages[sorted_dates[0]]]
        dates["J0"]  = sorted_dates[0]
    elif n == 2:
        groups["J0"] = [pg["text"] for _, pg in date_pages[sorted_dates[0]]]
        groups["J3"] = [pg["text"] for _, pg in date_pages[sorted_dates[1]]]
        dates["J0"]  = sorted_dates[0]
        dates["J3"]  = sorted_dates[1]
    else:
        j0_dt = parse_date(sorted_dates[0])
        groups["J0"] = [pg["text"] for _, pg in date_pages[sorted_dates[0]]]
        dates["J0"]  = sorted_dates[0]
        groups["J30"] = [pg["text"] for _, pg in date_pages[sorted_dates[-1]]]
        dates["J30"]  = sorted_dates[-1]

        j3_explicit_date = None
        j3_window_dates = []

        for d in sorted_dates[1:-1]:
            d_dt = parse_date(d)
            delta = abs((d_dt - j0_dt).days) if (j0_dt and d_dt) else 999
            # Explicit J3 label always wins — never merge into J0 even if within merge window
            if j3_explicit_date is None and _has_explicit_j3_label(date_pages[d]):
                j3_explicit_date = d
                groups["J3"].extend(pg["text"] for _, pg in date_pages[d])
            elif delta <= j0_merge_days:
                groups["J0"].extend(pg["text"] for _, pg in date_pages[d])
            else:
                if 2 <= delta <= 5:
                    j3_window_dates.append(d)
                groups["J3"].extend(pg["text"] for _, pg in date_pages[d])

        if j3_explicit_date:
            dates["J3"] = j3_explicit_date
        elif j3_window_dates:
            best = max(j3_window_dates, key=lambda d: _lab_density(date_pages[d]))
            dates["J3"] = best
        elif sorted_dates[1:-1]:
            for d in sorted_dates[1:-1]:
                d_dt = parse_date(d)
                delta = abs((d_dt - j0_dt).days) if (j0_dt and d_dt) else 999
                if delta > j0_merge_days:
                    dates["J3"] = d
                    break

    for _, pg in undated:
        groups["J0"].append(pg["text"])
    return {
        tp: {"text": "\n".join(groups[tp]), "date": dates[tp]}
        for tp in ["J0", "J3", "J30"]
    }

# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — Qwen Q&A extraction
# ─────────────────────────────────────────────────────────────────────────────
def get_cnr(cnr_sections: list, field: str) -> str:
    """Get a value from CNR form fields."""
    for sec in cnr_sections:
        v = sec["cnr_fields"].get(field, "").strip()
        if v and norm(v) not in ("none", "nsp", ""):
            return v
    return ""


def call_qwen(client: ollama.Client, model: str, prompt: str,
              retries: int = 2) -> str:
    """Call Qwen and return the response text. Retries on failure."""
    for attempt in range(retries + 1):
        try:
            response = client.chat(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                options={
                    "temperature": 0,
                    "top_p": 1.0,
                    "num_predict": 8192,
                }
            )
            return response["message"]["content"].strip()
        except Exception as e:
            if attempt < retries:
                print(f"    ⚠️  Qwen call failed (attempt {attempt+1}): {e}, retrying...")
                time.sleep(2)
            else:
                print(f"    ❌ Qwen call failed after {retries+1} attempts: {e}")
                return ""
    return ""


def parse_json_response(text: str) -> dict:
    """
    Extract JSON from Qwen response.
    Handles cases where Qwen adds explanation text around the JSON.
    """
    # Remove <think>...</think> blocks if present (Qwen3 thinking mode)
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()

    # Try direct parse first
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try to extract JSON block from markdown
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    # Try to find first { ... } block
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass

    print(f"    ⚠️  Could not parse JSON from response: {text[:200]}")
    return {}


def build_prompt(text_chunk: str, section_name: str,
                 fields: dict, schema: str, questions_text: str) -> str:
    return f"""You are a medical data extraction assistant. Extract information from this French pediatric malaria medical record.

SECTION: {section_name}

MEDICAL TEXT:
{text_chunk}

TASK: Answer these questions about the text above:
{questions_text}

RULES:
1. Answer ONLY from the text above — do not invent or assume values
2. For boolean fields: answer "Oui" if clearly present/positive, "Non" if clearly absent/negative, "" if not mentioned
3. For numeric fields: extract the number only (e.g. "38.5" not "38.5°C")
4. For dates: format as DD/MM/YYYY
5. "pas de X", "sans X", "absence de X" = "Non"; explicit mention of X = "Oui"
6. Return ONLY valid JSON, no explanation, no markdown

Return this exact JSON structure with your answers:
{schema}"""


def extract_section(client: ollama.Client, model: str,
                    text: str, section_name: str,
                    fields: dict) -> dict:
    """
    Extract all fields for a given section using Qwen Q&A.
    Chunks large texts and merges results — no truncation.

    fields: dict of {column_name: {question, type, ...}}
    Returns: dict of {column_name: value}
    """
    if not text.strip():
        return {}

    # Build the JSON schema
    schema_lines = []
    for col, cfg in fields.items():
        typ  = cfg.get("type", "text")
        hint = cfg.get("hint", "")
        if typ == "boolean":
            schema_lines.append(
                f'  "{col}": "<Oui or Non or empty string if not mentioned>"'
                + (f"  // {hint}" if hint else "")
            )
        elif typ == "numeric":
            schema_lines.append(
                f'  "{col}": "<number as string or empty string if not found>"'
                + (f"  // {hint}" if hint else "")
            )
        elif typ == "date":
            schema_lines.append(
                f'  "{col}": "<date as DD/MM/YYYY or empty string>"'
                + (f"  // {hint}" if hint else "")
            )
        else:
            schema_lines.append(
                f'  "{col}": "<value or empty string>"'
                + (f"  // {hint}" if hint else "")
            )

    schema = "{\n" + ",\n".join(schema_lines) + "\n}"
    questions_text = "\n".join(
        f"- {col}: {cfg.get('question', col)}"
        for col, cfg in fields.items()
    )

    # ── Batch extraction: small field groups + full text ─────────────────────
    # The prompt limit is caused by schema+questions size, not text size.
    # Solution: batch fields into groups of BATCH_SIZE, send full text each time.
    # First non-empty value wins per field across all batches.
    TEXT_CHARS = 14000   # safe text size — works reliably with small schemas
    BATCH_SIZE = 8       # fields per Qwen call — keeps prompt under limit

    field_items = list(fields.items())
    merged: dict[str, str] = {}

    for batch_start in range(0, len(field_items), BATCH_SIZE):
        batch = dict(field_items[batch_start:batch_start + BATCH_SIZE])

        # Build schema for this batch only
        batch_schema_lines = []
        for col, cfg in batch.items():
            typ  = cfg.get("type", "text")
            hint = cfg.get("hint", "")
            if typ == "boolean":
                batch_schema_lines.append(
                    f'  "{col}": "<Oui or Non or empty string>"' +
                    (f'  // {hint}' if hint else ""))
            elif typ == "numeric":
                batch_schema_lines.append(
                    f'  "{col}": "<number as string or empty>"' +
                    (f'  // {hint}' if hint else ""))
            elif typ == "date":
                batch_schema_lines.append(
                    f'  "{col}": "<DD/MM/YYYY or empty>"' +
                    (f'  // {hint}' if hint else ""))
            else:
                batch_schema_lines.append(
                    f'  "{col}": "<value or empty>"' +
                    (f'  // {hint}' if hint else ""))

        batch_schema = "{\n" + ",\n".join(batch_schema_lines) + "\n}"
        batch_questions = "\n".join(
            f"- {col}: {cfg.get('question', col)}"
            for col, cfg in batch.items()
        )

        # Send full text (first TEXT_CHARS) with this small batch
        prompt = build_prompt(
            text[:TEXT_CHARS], section_name, batch, batch_schema, batch_questions
        )
        raw = call_qwen(client, model, prompt)
        if not raw:
            continue

        result = parse_json_response(raw)
        for col, val in result.items():
            if col not in fields:
                continue
            val = str(val).strip() if val is not None else ""
            if val.lower() in ("null", "none", "n/a", "na", "-"):
                val = ""
            if val and not merged.get(col):
                merged[col] = val

        # Second pass on remainder for any still-empty fields in this batch
        still_empty = [c for c in batch if not merged.get(c)]
        if still_empty and len(text) > TEXT_CHARS:
            batch2 = {c: fields[c] for c in still_empty}
            batch2_schema_lines = []
            for col, cfg in batch2.items():
                typ = cfg.get("type", "text")
                if typ == "boolean":
                    batch2_schema_lines.append(f'  "{col}": "<Oui or Non or empty>"')
                elif typ == "numeric":
                    batch2_schema_lines.append(f'  "{col}": "<number or empty>"')
                else:
                    batch2_schema_lines.append(f'  "{col}": "<value or empty>"')
            batch2_schema = "{\n" + ",\n".join(batch2_schema_lines) + "\n}"
            batch2_questions = "\n".join(
                f"- {col}: {cfg.get('question', col)}"
                for col, cfg in batch2.items()
            )
            raw2 = call_qwen(client, model,
                             build_prompt(text[TEXT_CHARS-500:TEXT_CHARS*2],
                                          section_name, batch2,
                                          batch2_schema, batch2_questions))
            if raw2:
                result2 = parse_json_response(raw2)
                for col, val in result2.items():
                    if col not in fields:
                        continue
                    val = str(val).strip() if val is not None else ""
                    if val.lower() in ("null", "none", "n/a", "na", "-"):
                        val = ""
                    if val and not merged.get(col):
                        merged[col] = val

    return merged


# ─────────────────────────────────────────────────────────────────────────────
# Step 4 — Main extraction orchestrator
# ─────────────────────────────────────────────────────────────────────────────
def extract_row(patient_id: str, parsed: dict, groups: dict,
                fields_cfg: dict, client: ollama.Client,
                model: str) -> dict:
    """
    Extract all 130 variables for one patient using Qwen Q&A.
    Fields are grouped into sections for efficient batched calls.
    """
    row = {col: "" for col in COLUMNS}
    row["ID_Patient"] = patient_id

    all_cnr      = parsed["all_cnr"]
    clinical_text = parsed["clinical_text"]
    full_text     = parsed["full_text"]

    j0_text  = groups["J0"]["text"]
    j3_text  = groups["J3"]["text"]
    j30_text = groups["J30"]["text"]

    row["Date_PEC_J0"]  = groups["J0"]["date"]
    row["Date_PEC_J3"]  = groups["J3"]["date"]
    row["Date_PEC_J30"] = groups["J30"]["date"]

    # ── Section 1: Demographics & travel ─────────────────────────────────────
    print("    [1/6] Demographics & travel...")
    demo_fields = {k: v for k, v in fields_cfg.items()
                   if v.get("section") == "demographics"}
    if demo_fields:
        result = extract_section(client, model, clinical_text,
                                 "Demographics and Travel", demo_fields)
        for col, val in result.items():
            if val:
                if fields_cfg[col].get("type") == "date":
                    row[col] = fmt_date(val)
                else:
                    row[col] = val

    # CNR overrides for demographics
    row["Pays_de_naissance"] = row.get("Pays_de_naissance") or \
                               get_cnr(all_cnr, "Pays de naissance")
    row["date_premiers_symptomes"] = row.get("date_premiers_symptomes") or \
        fmt_date(get_cnr(all_cnr, "Date des Premiers Symptômes de cet accès") or
                 get_cnr(all_cnr, "Date des Premiers Symptomes de cet accès"))

    # Durée from date arithmetic fallback
    if not row.get("Durée_zone_endémie"):
        dep = parse_date(row.get("Date_départ_séjour", ""))
        ret = parse_date(row.get("Date_retour_séjour", ""))
        if dep and ret and ret > dep:
            days = (ret - dep).days
            if days >= 28:
                row["Durée_zone_endémie"] = f"{round(days/30)} mois"
            elif days >= 7:
                row["Durée_zone_endémie"] = f"{round(days/7)} semaines"
            else:
                row["Durée_zone_endémie"] = f"{days} jours"

    # ── Section 2: J0 — Admission ─────────────────────────────────────────────
    print("    [2/6] J0 — Admission...")
    j0_fields = {k: v for k, v in fields_cfg.items()
                 if v.get("section") == "J0"}
    if j0_fields and j0_text.strip():
        result = extract_section(client, model, j0_text, "J0 Admission", j0_fields)
        for col, val in result.items():
            if val:
                row[col] = val

    # ── Section 3: J0 Lab values ──────────────────────────────────────────────
    print("    [3/6] J0 Lab values...")
    j0_lab_fields = {k: v for k, v in fields_cfg.items()
                     if v.get("section") == "J0_lab"}
    if j0_lab_fields and j0_text.strip():
        result = extract_section(client, model, j0_text, "J0 Laboratory Results", j0_lab_fields)
        for col, val in result.items():
            if val:
                row[col] = val

    # ── Section 4: J3 ────────────────────────────────────────────────────────
    print("    [4/6] J3 — Follow-up...")
    j3_fields = {k: v for k, v in fields_cfg.items()
                 if v.get("section") in ("J3", "J3_lab")}
    if j3_fields and j3_text.strip():
        result = extract_section(client, model, j3_text, "J3 Follow-up", j3_fields)
        for col, val in result.items():
            if val:
                row[col] = val

    # ── Section 5: J30 ───────────────────────────────────────────────────────
    print("    [5/6] J30 — Outcome...")
    j30_fields = {k: v for k, v in fields_cfg.items()
                  if v.get("section") in ("J30", "J30_lab")}
    if j30_fields and j30_text.strip():
        result = extract_section(client, model, j30_text, "J30 Outcome", j30_fields)
        for col, val in result.items():
            if val:
                row[col] = val

    # J30 presence = PDV
    if groups["J30"]["text"].strip():
        row["PDV_J30"] = "Oui"

    # fièvre_J30 default: if J30 text has no clinical note → Non
    if not row.get("fièvre_J30") and groups["J30"]["text"].strip():
        j30_clinical_keywords = ["apyretique", "apyrexie", "fievre", "febrile",
                                  "temperature", "temp", "examen clinique",
                                  "etat general", "beg", "va bien"]
        import unicodedata
        j30_norm = unicodedata.normalize("NFKD", groups["J30"]["text"].lower())
        j30_norm = "".join(c for c in j30_norm if not unicodedata.combining(c))
        has_clinical = any(kw in j30_norm for kw in j30_clinical_keywords)
        if not has_clinical:
            row["fièvre_J30"] = "Non"  # lab-only J30, no clinical note

    # ── Section 6: Treatment ──────────────────────────────────────────────────
    print("    [6/6] Treatment & outcome...")
    treat_fields = {k: v for k, v in fields_cfg.items()
                    if v.get("section") == "treatment"}
    if treat_fields:
        result = extract_section(client, model, clinical_text,
                                 "Treatment and Outcome", treat_fields)
        for col, val in result.items():
            if val:
                row[col] = val

    # ── CNR overrides (always trust CNR for these fields) ─────────────────────
    # Type paludisme — CNR is authoritative
    cnr_species = get_cnr(all_cnr, "Espèce(s) Plasmodiale(s)") or \
                  get_cnr(all_cnr, "Espece(s) Plasmodiale(s)")
    if cnr_species:
        species_map = {
            "falciparum": "P. falciparum",
            "ovale":      "P. ovale",
            "vivax":      "P. vivax",
            "malariae":   "P. malariae",
            "knowlesi":   "P. knowlesi",
            "spp":        "Plasmodium spp",
        }
        for key, val in species_map.items():
            if key in norm(cnr_species):
                row["Type_paludisme"] = val
                break

    # Hospitalisation from CNR
    if not row["Hospitalisation"]:
        row["Hospitalisation"] = get_cnr(all_cnr, "Hospitalisation")

    # Durée hospit from CNR
    if not row["Durée_hospit"]:
        row["Durée_hospit"] = get_cnr(all_cnr, "Nombre de jours d'hospitalisation")

    # Durée traitement from CNR
    if not row["Durée_traitement"]:
        row["Durée_traitement"] = get_cnr(all_cnr, "Durée en jours")

    # HRP2 from CNR — overrides Qwen (result split across pages causes misses)
    if not row["HRP2_J0"]:
        for sec in all_cnr:
            hrp2_raw = (sec["cnr_fields"].get("Bandelettes résultat", "") or
                        sec["cnr_fields"].get("Bandelettes resultat", "") or
                        sec["cnr_fields"].get("Bandelettes (HRP2, LDH, ...)", ""))
            if hrp2_raw:
                hrp2_norm = norm(hrp2_raw)
                if "positif" in hrp2_norm or "positive" in hrp2_norm:
                    row["HRP2_J0"] = "Oui"
                elif "negatif" in hrp2_norm or "negative" in hrp2_norm:
                    row["HRP2_J0"] = "Non"
                if row["HRP2_J0"]:
                    break

    # Température_J3 from CNR — only available when no J3 clinical visit
    if not row.get("Température_J3"):
        for sec in all_cnr:
            t = sec["cnr_fields"].get("J3 ou J4 Température", "")
            if t:
                row["Température_J3"] = t.strip()
                break

    # Frottis_sanguin_J3 from CNR when empty
    if not row.get("Frottis_sanguin_J3"):
        for sec in all_cnr:
            p = norm(sec["cnr_fields"].get("J3 ou J4 Parasitologie", ""))
            if p:
                row["Frottis_sanguin_J3"] = "Non" if "absence" in p else "Oui"
                break

    # Goutte épaisse from CNR (bypasses NSP filter)
    for sec in all_cnr:
        raw = (sec["cnr_fields"].get("Goutte épaisse", "") or
               sec["cnr_fields"].get("Goutte epaisse", ""))
        if raw.strip():
            row["Goutte_épaisse_J0"] = raw.strip()
            break

    # palu_ant from CNR
    if not row["palu_ant"]:
        for sec in all_cnr:
            raw = (sec["cnr_fields"].get(
                "Antécédents de paludisme dans les 3 derniers mois", "") or
                   sec["cnr_fields"].get(
                "Antecedents de paludisme dans les 3 derniers mois", ""))
            if raw.strip() and norm(raw) == "oui":
                row["palu_ant"] = "Oui"
                break

    return row


# ─────────────────────────────────────────────────────────────────────────────
# Step 5 — Write outputs (same as before)
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(rows: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️  openpyxl not found: pip install openpyxl --break-system-packages")
        return

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Données patients"
        hdr  = PatternFill("solid", fgColor="FFD966")
        col  = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="AAAAAA")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        sections_def = [
            ("Profil épidémiologique", 1, 21),
            ("J0 – Admission",        22, 62),
            ("J3 – Contrôle",         63, 103),
            ("J30 – Suivi",           104, 144),
            ("Traitement & issue",    145, len(COLUMNS)),
        ]
        for title, s, e in sections_def:
            e = min(e, len(COLUMNS))
            if s > len(COLUMNS): break
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
            c = ws.cell(row=1, column=s)
            c.value = title
            c.font = Font(bold=True, size=10)
            c.fill = hdr
            c.alignment = Alignment(horizontal="center", vertical="center")

        for ci, cname in enumerate(COLUMNS, 1):
            c = ws.cell(row=2, column=ci)
            c.value = cname
            c.font = Font(bold=True, size=9)
            c.fill = col
            c.border = brd
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        ws.freeze_panes = "A3"
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 42
        for ci in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14
        ws.column_dimensions["A"].width = 18

    for row_data in rows:
        pid = row_data["ID_Patient"]
        target = None
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value == pid:
                target = r
                break
        if target is None:
            target = max(ws.max_row + 1, 3)
        for ci, cname in enumerate(COLUMNS, 1):
            c = ws.cell(row=target, column=ci)
            c.value = row_data.get(cname, "")
            c.alignment = Alignment(vertical="top", wrap_text=False)

    wb.save(path)
    print(f"  ✅ Excel: {path}")


def write_sqlite(rows: list[dict], path: Path) -> None:
    conn = sqlite3.connect(path)
    cur  = conn.cursor()
    cols_sql = ", ".join(f'"{c}" TEXT' for c in COLUMNS)
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS patients (
            {cols_sql},
            extraction_date TEXT,
            PRIMARY KEY ("ID_Patient")
        )
    """)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for rd in rows:
        col_names    = ", ".join(f'"{c}"' for c in COLUMNS) + ', "extraction_date"'
        placeholders = ", ".join("?" for _ in range(len(COLUMNS) + 1))
        values       = [rd.get(c, "") for c in COLUMNS] + [ts]
        cur.execute(
            f"INSERT OR REPLACE INTO patients ({col_names}) VALUES ({placeholders})",
            values
        )
    conn.commit()
    conn.close()
    print(f"  ✅ SQLite: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def process_patient(anon_file: Path, fields_cfg: dict,
                    client: ollama.Client, model: str) -> Optional[dict]:
    m = re.search(r"patient_(\d+)_anonymized", anon_file.name, re.IGNORECASE)
    if not m:
        print(f"⚠️  Cannot parse patient ID from: {anon_file.name}")
        return None

    patient_id = f"PATIENT_{m.group(1).zfill(3)}"
    print(f"\n{'='*60}")
    print(f"  Patient: {patient_id}  ({anon_file.name})")

    t0 = time.time()

    parsed = parse_anonymized_file(anon_file)
    groups = assign_timepoints(parsed)

    print(f"  Timepoints: "
          f"J0={groups['J0']['date'] or '—'}  "
          f"J3={groups['J3']['date'] or '—'}  "
          f"J30={groups['J30']['date'] or '—'}")

    row = extract_row(patient_id, parsed, groups, fields_cfg, client, model)

    elapsed = time.time() - t0
    print(f"\n  Extracted in {elapsed:.1f}s. Key values:")

    show = ["Sexe","Age","Lieu_séjour","chimioprophylaxie","moustiquaire_imp",
            "Type_paludisme","gravité_palu","Traitement_antipalu",
            "Hospitalisation","Durée_hospit",
            "Date_PEC_J0","Poids_J0","Température_J0","Fréquence_Cardiaque_J0",
            "hémoglobine_J0","plaquettes_J0","ASAT_J0","CRP_J0",
            "Parasitémie_J0","Frottis_sanguin_J0","HRP2_J0",
            "Date_PEC_J3","fièvre_J3","Frottis_sanguin_J3",
            "Date_PEC_J30","PDV_J30","fièvre_J30","Frottis_sanguin_J30"]

    for col in show:
        val = row.get(col, "")
        status = "✅" if val else "❌"
        print(f"    {status} {col}: {val or '(empty)'}")

    return row


def main():
    parser = argparse.ArgumentParser(
        description="System B Qwen Q&A variable extraction → Excel + SQLite"
    )
    grp = parser.add_mutually_exclusive_group(required=True)
    grp.add_argument("--patient", help="Path to patient_NNN_anonymized.txt")
    grp.add_argument("--all", action="store_true")
    parser.add_argument("--fields",     default=str(FIELDS_PATH))
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR))
    parser.add_argument("--model",      default=DEFAULT_MODEL)
    parser.add_argument("--host",       default=OLLAMA_HOST)
    args = parser.parse_args()

    print(f"  Model: {args.model}")
    print(f"  Host:  {args.host}")

    # Load field definitions
    fields_cfg = load_fields(Path(args.fields))
    if not fields_cfg:
        print("❌ No fields configuration found. Create extraction_fields.yaml first.")
        sys.exit(1)
    print(f"  Fields: {len(fields_cfg)} variables configured")

    # Connect to Qwen
    client = ollama.Client(host=args.host)

    # Warm up — ensure model is loaded in VRAM
    print("  Warming up model...")
    t0 = time.time()
    client.chat(
        model=args.model,
        messages=[{"role": "user", "content": "Ready?"}],
        options={"temperature": 0, "num_predict": 5}
    )
    print(f"  Model ready in {time.time()-t0:.1f}s")

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.patient:
        files = [Path(args.patient)]
    else:
        files = sorted(FINAL_ANON.glob("patient_*_anonymized.txt"))
        if not files:
            print(f"No files found in {FINAL_ANON}")
            sys.exit(1)
        print(f"Found {len(files)} patient file(s)")

    rows = []
    for f in files:
        row = process_patient(f, fields_cfg, client, args.model)
        if row:
            rows.append(row)

    if rows:
        write_excel(rows, out_dir / "research_table_qwen.xlsx")
        write_sqlite(rows, out_dir / "research_database_qwen.db")
        print(f"\n{'='*60}")
        print(f"✅ Done: {len(rows)} patient(s) processed")
    else:
        print("No rows extracted.")
        sys.exit(1)


if __name__ == "__main__":
    main()