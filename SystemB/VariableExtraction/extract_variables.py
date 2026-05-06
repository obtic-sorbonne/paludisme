#!/usr/bin/env python3
"""
extract_variables.py  –  System B Variable Extraction
Location: ~/digitize_medical_records/VariableExtraction/extract_variables.py

Reads patient_NNN_anonymized.txt  →  130 clinical variables  →  Excel + SQLite

Usage:
  python extract_variables.py --patient patient_002_anonymized.txt
  python extract_variables.py --all
  python extract_variables.py --all --config /path/to/other_config.yaml

New institution: only edit variable_extraction_config.yaml
  - institution.services
  - lieu_sejour_keywords
  - chimio_keywords
  - treatment_keywords

J0/J3/J30 rule (consultation ORDER, not date arithmetic):
  Collect all unique explicit page-level dates, sort chronologically.
  1 date  → J0 only
  2 dates → J0, J3
  3+ dates:
    First = J0
    Last  = J30
    Middle dates ≤ j0_merge_days from J0 → merged into J0
    First middle date beyond threshold   → J3
  Never derives dates from ranges like "DU X AU Y"

Priority: clinical document > CNR form  (configurable per field)
"""
from __future__ import annotations

import argparse
import re
import sqlite3
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml

# ── Paths ──────────────────────────────────────────────────────────────────────
WORK_DIR    = Path("/home/lfarooq/digitize_medical_records")
FINAL_ANON  = WORK_DIR / "outputs" / "final_anonymized"
OUTPUT_DIR  = WORK_DIR / "VariableExtraction" / "outputs"
CONFIG_PATH = WORK_DIR / "VariableExtraction" / "variable_extraction_config.yaml"

# ── Column list (130 variables) ────────────────────────────────────────────────
COLUMNS = [
    "ID_Patient","Sexe","Age","Pays_de_naissance","Durée_zone_endémie",
    "Date_départ_séjour","Date_retour_séjour","Lieu_séjour","Symptômes_séjour",
    "type_symptômes_séjour","Traitement_séjour","Drépanocytose","Splénectomie",
    "palu_ant","diabète","insuff_rénale","chimioprophylaxie","moustiquaire_imp",
    "consultation_pré_voyage","date_premiers_symptomes","méd_traitant",
    "Date_PEC_J0","fièvre_J0","anorexie_J0","diarrhées_J0","vomissement_J0",
    "nausée_J0","douleurs_abdominales_J0","asthénie_J0","trouble_conscience_J0",
    "convulsion_J0","céphalées_J0","autres_motifs_J0","Poids_J0","Température_J0",
    "Fréquence_Cardiaque_J0","SaO2_J0","Hépatomégalie_J0","spénomégalie_J0",
    "pâleur_J0","ictère_J0","raideur_nuque_J0","Goutte_épaisse_J0",
    "Frottis_sanguin_J0","Parasitémie_J0","PCR_J0","HRP2_J0","hémoculture_J0",
    "hémoglobine_J0","plaquettes_J0","globules_blancs_J0","eosinophiles_J0",
    "ASAT_J0","ALAT_J0","bilirubine_J0","Urée_J0","Créatinine_J0","CRP_J0",
    "PCT_J0","pH_J0","lactates_J0","TP_J0","TCA_J0",
    "Date_PEC_J3","fièvre_J3","anorexie_J3","diarrhées_J3","vomissement_J3",
    "nausée_J3","douleurs_abdominales_J3","asthénie_J3","trouble_conscience_J3",
    "convulsion_J3","céphalées_J3","autres_motifs_J3","Poids_J3","Température_J3",
    "Fréquence_Cardiaque_J3","SaO2_J3","Hépatomégalie_J3","spénomégalie_J3",
    "pâleur_J3","ictère_J3","raideur_nuque_J3","Goutte_épaisse_J3",
    "Frottis_sanguin_J3","Parasitémie_J3","PCR_J3","HRP2_J3","hémoculture_J3",
    "hémoglobine_J3","plaquettes_J3","globules_blancs_J3","eosinophiles_J3",
    "ASAT_J3","ALAT_J3","bilirubine_J3","Urée_J3","Créatinine_J3","CRP_J3",
    "PCT_J3","pH_J3","lactates_J3","TP_J3","TCA_J3",
    "PDV_J30","Date_PEC_J30","fièvre_J30","anorexie_J30","diarrhées_J30",
    "vomissement_J30","nausée_J30","douleurs_abdominales_J30","asthénie_J30",
    "trouble_conscience_J30","convulsion_J30","céphalées_J30","autres_motifs_J30",
    "Poids_J30","Température_J30","Fréquence_Cardiaque_J30","SaO2_J30",
    "Hépatomégalie_J30","spénomégalie_J30","pâleur_J30","ictère_J30",
    "raideur_nuque_J30","Goutte_épaisse_J30","Frottis_sanguin_J30",
    "Parasitémie_J30","PCR_J30","HRP2_J30","hémoculture_J30","hémoglobine_J30",
    "plaquettes_J30","globules_blancs_J30","eosinophiles_J30","ASAT_J30",
    "ALAT_J30","bilirubine_J30","Urée_J30","Créatinine_J30","CRP_J30",
    "PCT_J30","pH_J30","lactates_J30","TP_J30","TCA_J30",
    "Type_paludisme","gravité_palu","Traitement_antipalu","Posologie_traitement",
    "Durée_traitement","Perfusion","Hospitalisation","Durée_hospit",
    "type_service","ECG",
]


# ─────────────────────────────────────────────────────────────────────────────
# Utilities
# ─────────────────────────────────────────────────────────────────────────────
def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def parse_date(s: str) -> Optional[datetime]:
    s = str(s).strip().replace(".", "/")
    m = re.match(r"(\d{1,2})\s+(\d{2})\s+(\d{4})", s)
    if m:
        s = f"{m.group(1).zfill(2)}/{m.group(2)}/{m.group(3)}"
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def fmt_date(s: str) -> str:
    d = parse_date(s)
    return d.strftime("%d/%m/%Y") if d else s.strip()


def first_match(patterns: list, text: str, flags: int = re.IGNORECASE) -> str:
    for pat in patterns:
        try:
            m = re.search(pat, text, flags)
            if m:
                return m.group(1).strip()
        except (re.error, IndexError):
            continue
    return ""


def convert_hb(value: str) -> str:
    """Convert Hb from g/l → g/dl when value ≥ 30."""
    try:
        v = float(str(value).replace(",", "."))
        return f"{v/10:.1f}" if v >= 30 else str(value)
    except ValueError:
        return str(value)


def load_config(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 – Parse anonymized file
# Each DocumentSection owns its pages; each page has its own explicit date.
# Sections with CNR fields also carry those fields.
# ─────────────────────────────────────────────────────────────────────────────

# Strong date patterns that mark a consultation / lab event date
_PRIMARY_DATE_PATTERNS = [
    r"Date\s*:\s*(\d{2}/\d{2}/\d{4})",
    r"Date\s*/\s*Heure\s+(\d{2}/\d{2}/\d{4})",
    r"Prélèvement\s*:\s*(\d{2}/\d{2}/\d{4})",
    r"Entrée Urgence le\s*(\d{2}/\d{2}/\d{4})",
    r"COMPTE RENDU DE CONSULTATION DU\s*(\d{2}/\d{2}/\d{2,4})",
    r"PARAMETRES A L[''']ARRIVEE\s*:\s*(\d{2}/\d{2}/\d{4})",
]
# Labels whose date should be IGNORED (birth date, travel dates, etc.)
_IGNORE_CTX = [
    "date naissance", "naiss", "date de depart", "date retour",
    "date des premiers", "date de naissance", "du diagnostic",
    "date de la derniere", "date de la premiere",
]


def _page_date(page_text: str) -> str:
    """Return the first explicit clinical date found in a page, or ''."""
    for pat in _PRIMARY_DATE_PATTERNS:
        for m in re.finditer(pat, page_text, re.IGNORECASE):
            ctx = norm(page_text[max(0, m.start()-60):m.start()])
            if any(ig in ctx for ig in _IGNORE_CTX):
                continue
            d = parse_date(m.group(1))
            if d and 2000 <= d.year <= 2035:
                return d.strftime("%d/%m/%Y")
    return ""


class Page:
    def __init__(self, num: int, text: str):
        self.num  = num
        self.text = text
        self.date = _page_date(text)


class DocumentSection:
    def __init__(self, doc_id: str, doc_type: str):
        self.doc_id    = doc_id
        self.doc_type  = doc_type
        self.pages: list[Page]          = []
        self.cnr_fields: dict[str, str] = {}
        self.is_clinical = False
        self.is_lab      = False
        self.is_cnr      = False

    def pages_for_date(self, date_str: str) -> list[Page]:
        return [p for p in self.pages if p.date == date_str]

    def text_for_date(self, date_str: str) -> str:
        pages = self.pages_for_date(date_str)
        # If no specific page matches, return full text (e.g. CNR form)
        if not pages and not date_str:
            return self.full_text
        return "\n".join(p.text for p in pages)

    @property
    def full_text(self) -> str:
        return "\n".join(p.text for p in self.pages)


def parse_anonymized_file(path: Path, config: dict) -> list[DocumentSection]:
    text  = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()

    sections: list[DocumentSection] = []
    current: Optional[DocumentSection] = None
    current_page_num: Optional[int] = None
    in_cnr = False
    page_lines: list[str] = []

    doc_re  = re.compile(r"^DOCUMENT:\s+(\S+)\s+\[(\w+)\]$")
    page_re = re.compile(r"^===== page-(\d+) =====$")

    def flush():
        if current and current_page_num is not None:
            current.pages.append(Page(current_page_num, "\n".join(page_lines)))

    for line in lines:
        # New document
        m = doc_re.match(line.strip())
        if m:
            flush(); page_lines.clear(); current_page_num = None; in_cnr = False
            current = DocumentSection(m.group(1), m.group(2))
            sections.append(current)
            continue

        if current is None:
            continue

        # CNR fields section
        if "── CNR FORM FIELDS" in line:
            flush(); page_lines.clear(); current_page_num = None
            in_cnr = True; current.is_cnr = True
            continue
        if "── CLINICAL / LAB DOCUMENTS" in line:
            in_cnr = False; continue

        if in_cnr:
            if ":" in line:
                k, _, v = line.partition(":")
                k, v = k.strip(), v.strip()
                if k and v and v.lower() != "none":
                    current.cnr_fields[k] = v
            continue

        # New page
        mp = page_re.match(line.strip())
        if mp:
            flush(); page_lines.clear()
            current_page_num = int(mp.group(1))
            continue

        if current_page_num is not None:
            page_lines.append(line)

    flush()

    # Classify each section
    ck = config.get("doc_type_keywords", {})
    for sec in sections:
        t = norm(sec.full_text)
        sec.is_clinical = any(norm(k) in t for k in ck.get("clinical", []))
        sec.is_lab      = any(norm(k) in t for k in ck.get("lab", []))

    return sections


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 – Assign J0 / J3 / J30
# Works at PAGE level so a mixed document can contribute pages to different TPs
# ─────────────────────────────────────────────────────────────────────────────
class TimeGroup:
    """All pages across all sections that share a clinical date."""
    def __init__(self, date_str: str):
        self.date = date_str
        self.pages: list[tuple[DocumentSection, Page]] = []   # (section, page)

    def text(self) -> str:
        return "\n".join(p.text for _, p in self.pages)

    def clinical_text(self) -> str:
        return "\n".join(p.text for s, p in self.pages if s.is_clinical)

    def lab_text(self) -> str:
        return "\n".join(p.text for s, p in self.pages if s.is_lab)

    def cnr_sections(self) -> list[DocumentSection]:
        seen, out = set(), []
        for s, _ in self.pages:
            if s.is_cnr and s.doc_id not in seen:
                seen.add(s.doc_id); out.append(s)
        return out


def assign_timepoints(
    sections: list[DocumentSection],
    j0_merge_days: int = 3,
) -> tuple[dict[str, TimeGroup], dict[str, str]]:
    """
    Returns:
      groups  – {"J0": TimeGroup, "J3": TimeGroup, "J30": TimeGroup}
      tp_dates – {"J0": "21/09/2006", ...}
    """
    # Collect all dated pages
    date_map: dict[str, TimeGroup] = {}
    undated_pages: list[tuple[DocumentSection, Page]] = []

    for sec in sections:
        for pg in sec.pages:
            if pg.date:
                if pg.date not in date_map:
                    date_map[pg.date] = TimeGroup(pg.date)
                date_map[pg.date].pages.append((sec, pg))
            else:
                undated_pages.append((sec, pg))

    sorted_dates = sorted(date_map.keys(),
                          key=lambda d: parse_date(d) or datetime.min)

    empty = TimeGroup("")
    groups   = {"J0": empty, "J3": empty, "J30": empty}
    tp_dates = {"J0": "",    "J3": "",    "J30": ""}

    n = len(sorted_dates)

    if n == 0:
        # No dates at all → everything J0
        g = TimeGroup("")
        for sec in sections:
            for pg in sec.pages:
                g.pages.append((sec, pg))
        groups["J0"] = g

    elif n == 1:
        groups["J0"]   = date_map[sorted_dates[0]]
        tp_dates["J0"] = sorted_dates[0]

    elif n == 2:
        groups["J0"]   = date_map[sorted_dates[0]]; tp_dates["J0"] = sorted_dates[0]
        groups["J3"]   = date_map[sorted_dates[1]]; tp_dates["J3"] = sorted_dates[1]

    else:  # 3+ dates
        j0_date  = sorted_dates[0]
        j30_date = sorted_dates[-1]
        j0_dt    = parse_date(j0_date)

        groups["J0"]    = date_map[j0_date];  tp_dates["J0"]  = j0_date
        groups["J30"]   = date_map[j30_date]; tp_dates["J30"] = j30_date

        j3_group   = TimeGroup("")
        j3_assigned = False

        for mid_date in sorted_dates[1:-1]:
            mid_dt     = parse_date(mid_date)
            days_delta = abs((mid_dt - j0_dt).days) if (j0_dt and mid_dt) else 999

            if days_delta <= j0_merge_days:
                # Merge into J0
                groups["J0"].pages.extend(date_map[mid_date].pages)
            else:
                # Goes to J3
                j3_group.pages.extend(date_map[mid_date].pages)
                if not j3_assigned:
                    j3_group.date  = mid_date
                    tp_dates["J3"] = mid_date
                    j3_assigned    = True

        groups["J3"] = j3_group

    # Undated pages → J0
    for item in undated_pages:
        groups["J0"].pages.append(item)

    # Also attach CNR fields to all groups for lookup
    return groups, tp_dates


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 – Extract variables
# ─────────────────────────────────────────────────────────────────────────────
def get_cnr(sections: list[DocumentSection], field: str) -> str:
    for sec in sections:
        v = sec.cnr_fields.get(field, "").strip()
        if v and v.lower() not in ("none", "nsp", ""):
            return v
    return ""


def detect_boolean(cfg: dict, text: str) -> str:
    ftype = cfg.get("type", "")
    t     = norm(text)

    if ftype == "boolean_presence":
        for pat in cfg.get("patterns", []):
            try:
                if re.search(norm(pat), t):
                    return "Oui"
            except re.error:
                pass
        return ""

    if ftype == "boolean_with_negative":
        # IMPORTANT: check FULL text for negative, then positive
        # Use the original text (not normed) for case-sensitive patterns
        for pat in cfg.get("patterns_negative", []):
            try:
                if re.search(pat, text, re.IGNORECASE):
                    return "Non"
            except re.error:
                pass
        for pat in cfg.get("patterns_positive", []):
            try:
                if re.search(pat, text, re.IGNORECASE):
                    return "Oui"
            except re.error:
                pass
    return ""


def detect_frottis(text: str) -> str:
    """
    Special handling for frottis: a document can contain BOTH a positive
    frottis result AND a later negative result (e.g. J0 positive, J3 negative
    on separate pages of DOC_00119).
    We look for BOTH and return whichever is more specific to the page context.
    Strategy: positive wins if it appears BEFORE any negative in the text.
    """
    pos_pats = [
        r"presence de trophozoites",
        r"présence de trophozoïtes",
        r"parasitemie\s*:\s*\d",
        r"parasitémie\s*à\s*\d",
    ]
    neg_pats = [
        r"il n'a pas ete vu de parasites",
        r"il n.a pas ete vu de parasites",
        r"pas de parasites intra ou extra",
        r"absence de parasites",
    ]

    pos_pos = min(
        (m.start() for pat in pos_pats
         for m in [re.search(pat, text, re.IGNORECASE)] if m),
        default=None
    )
    neg_pos = min(
        (m.start() for pat in neg_pats
         for m in [re.search(pat, text, re.IGNORECASE)] if m),
        default=None
    )

    if pos_pos is None and neg_pos is None:
        return ""
    if pos_pos is not None and neg_pos is None:
        return "Oui"
    if neg_pos is not None and pos_pos is None:
        return "Non"
    # Both present → whichever appears first wins
    return "Oui" if pos_pos < neg_pos else "Non"


def extract_hb(text: str, all_cnr: list, prefix: str) -> str:
    m = re.search(r"Hemoglobine\.+\s*([\d,\.]+)", text, re.IGNORECASE)
    if m:
        return m.group(1).replace(",", ".")
    m = re.search(r"Hémoglobine\s*\(g/dl\)\s*\|\s*([\d,\.]+)", text, re.IGNORECASE)
    if m:
        return m.group(1).replace(",", ".")
    if prefix == "J0":
        v = get_cnr(all_cnr, "Hémoglobine (g/l)")
        return convert_hb(v) if v else ""
    return ""


def extract_platelets(text: str, all_cnr: list, prefix: str) -> str:
    # FIX: stop at first non-digit/comma/dot after the number
    m = re.search(r"PLAQUETTES\.+\s*([\d,\.]+)(?:\s|$)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().replace(",", ".")
    if prefix == "J0":
        return get_cnr(all_cnr, "Plaquettes (giga/l)")
    return ""


def extract_wbc(text: str, all_cnr: list, prefix: str) -> str:
    m = re.search(r"LEUCOCYTES\.+\s*([\d,\.]+)", text, re.IGNORECASE)
    if m:
        return m.group(1).replace(",", ".")
    if prefix == "J0":
        return get_cnr(all_cnr, "GB (giga/l)")
    return ""


def detect_severity(cfg: dict, clinical_text: str, all_cnr: list) -> str:
    t = norm(clinical_text)
    for pat in cfg.get("patterns_severe", []):
        if re.search(norm(pat), t):
            return "grave"
    for pat in cfg.get("patterns_simple", []):
        if re.search(norm(pat), t):
            return "simple"
    cnr_val = get_cnr(all_cnr, cfg.get("cnr_field", ""))
    if cnr_val:
        nv = norm(cnr_val)
        if "grave" in nv: return "grave"
        if "simple" in nv or "sans vomissement" in nv: return "simple"
    return ""


def fill_timepoint(
    row: dict,
    prefix: str,
    grp: TimeGroup,
    all_cnr: list[DocumentSection],
    fcfg: dict,
    tp_date: str,
) -> None:
    """Fill all J{prefix} fields from the TimeGroup."""
    clinical = grp.clinical_text()
    lab      = grp.lab_text()
    all_text = grp.text()

    row[f"Date_PEC_{prefix}"] = tp_date

    # ── Vitals ────────────────────────────────────────────────────────────────
    vital_map = {
        f"Poids_{prefix}":              "Poids_J0",
        f"Température_{prefix}":        "Température_J0",
        f"Fréquence_Cardiaque_{prefix}":"Fréquence_Cardiaque_J0",
        f"SaO2_{prefix}":               "SaO2_J0",
    }
    for col, cfg_key in vital_map.items():
        if col in COLUMNS and not row[col]:
            pats = fcfg.get(cfg_key, {}).get("patterns", [])
            val  = first_match(pats, clinical + "\n" + all_text)
            if val:
                row[col] = val.replace(",", ".")

    # ── Boolean symptoms ──────────────────────────────────────────────────────
    for base in ["fièvre","anorexie","diarrhées","vomissement","nausée",
                 "douleurs_abdominales","asthénie","trouble_conscience",
                 "convulsion","céphalées","pâleur","ictère"]:
        col = f"{base}_{prefix}"
        if col in COLUMNS and not row[col]:
            cfg = fcfg.get(f"{base}_J0", {})
            if cfg:
                row[col] = detect_boolean(cfg, clinical + "\n" + all_text)

    # ── Exam findings ─────────────────────────────────────────────────────────
    for base in ["Hépatomégalie","spénomégalie","raideur_nuque"]:
        col = f"{base}_{prefix}"
        if col in COLUMNS and not row[col]:
            cfg = fcfg.get(f"{base}_J0", {})
            if cfg:
                row[col] = detect_boolean(cfg, clinical)

    # ── Lab values ────────────────────────────────────────────────────────────
    full_lab = lab + "\n" + all_text
    if not row[f"hémoglobine_{prefix}"]:
        row[f"hémoglobine_{prefix}"] = extract_hb(full_lab, all_cnr, prefix)
    if not row[f"plaquettes_{prefix}"]:
        row[f"plaquettes_{prefix}"] = extract_platelets(full_lab, all_cnr, prefix)
    if not row[f"globules_blancs_{prefix}"]:
        row[f"globules_blancs_{prefix}"] = extract_wbc(full_lab, all_cnr, prefix)

    for base in ["ASAT","ALAT","bilirubine","Urée","Créatinine","CRP"]:
        col = f"{base}_{prefix}"
        if col in COLUMNS and not row[col]:
            pats = fcfg.get(f"{base}_J0", {}).get("patterns", [])
            val  = first_match(pats, full_lab)
            if val:
                row[col] = val.rstrip("+").strip()

    # ── Parasitology ──────────────────────────────────────────────────────────
    if not row[f"Parasitémie_{prefix}"]:
        pats = fcfg.get("Parasitémie_J0", {}).get("patterns", [])
        val  = first_match(pats, all_text)
        if not val and prefix == "J0":
            val = get_cnr(all_cnr, "Densité parasitaire")
        if val:
            row[f"Parasitémie_{prefix}"] = val.replace(",", ".")

    # Frottis: use special detection that respects order
    if not row[f"Frottis_sanguin_{prefix}"]:
        row[f"Frottis_sanguin_{prefix}"] = detect_frottis(all_text)

    if not row[f"HRP2_{prefix}"]:
        cfg = fcfg.get("HRP2_J0", {})
        if cfg:
            row[f"HRP2_{prefix}"] = detect_boolean(cfg, all_text)

    # ── J3 specific ───────────────────────────────────────────────────────────
    if prefix == "J3":
        # CNR J3/J4 temperature takes precedence
        cnr_temp = get_cnr(all_cnr, "J3 ou J4 Température")
        if cnr_temp:
            row["Température_J3"] = cnr_temp

        if not row["Frottis_sanguin_J3"]:
            v = get_cnr(all_cnr, "J3 ou J4 Parasitologie")
            if v:
                row["Frottis_sanguin_J3"] = "Non" if "absence" in norm(v) else "Oui"

    # ── J30 specific ──────────────────────────────────────────────────────────
    if prefix == "J30":
        if grp.pages:
            row["PDV_J30"] = "Oui"
        if not row["Poids_J30"]:
            m = re.search(r"Poids\s+([\d,\.]+)\s*kg", all_text, re.IGNORECASE)
            if m:
                row["Poids_J30"] = m.group(1)
        # J30 frottis: "bilan normal" → negative
        if not row["Frottis_sanguin_J30"]:
            if re.search(r"bilan h.matologique normal|bilan normal", all_text, re.IGNORECASE):
                row["Frottis_sanguin_J30"] = "Non"


def extract_row(
    patient_id: str,
    sections: list[DocumentSection],
    groups: dict[str, TimeGroup],
    tp_dates: dict[str, str],
    config: dict,
) -> dict[str, str]:
    row   = {col: "" for col in COLUMNS}
    row["ID_Patient"] = patient_id
    fcfg  = config.get("fields", {})

    all_cnr       = [s for s in sections if s.cnr_fields]
    all_clinical  = [s for s in sections if s.is_clinical]
    clinical_text = "\n".join(s.full_text for s in all_clinical)
    all_text      = "\n".join(s.full_text for s in sections)

    # ── Demographics ──────────────────────────────────────────────────────────
    row["Sexe"] = (
        first_match(fcfg.get("Sexe",{}).get("patterns",[]), clinical_text)
        or get_cnr(all_cnr, "Sexe")
    )
    age = first_match(fcfg.get("Age",{}).get("patterns",[]), clinical_text)
    row["Age"] = age.rstrip("a").strip() if age else ""
    row["Pays_de_naissance"] = get_cnr(all_cnr, "Pays de naissance")

    # ── Travel ────────────────────────────────────────────────────────────────
    def get_travel(cfg_key, cnr_key):
        v = first_match(fcfg.get(cfg_key,{}).get("patterns",[]), clinical_text)
        if not v:
            v = get_cnr(all_cnr, cnr_key)
        return fmt_date(v) if v else ""

    row["Date_départ_séjour"] = get_travel("Date_départ_séjour", "Date départ")
    row["Date_retour_séjour"] = get_travel("Date_retour_séjour", "Date retour")
    row["Durée_zone_endémie"] = first_match(
        fcfg.get("Durée_zone_endémie",{}).get("patterns",[]), clinical_text
    )

    # Lieu séjour from keyword map
    t = norm(all_text)
    for kw, value in config.get("lieu_sejour_keywords", {}).items():
        if kw in t:
            row["Lieu_séjour"] = value
            break

    # Symptom onset
    v = get_cnr(all_cnr, "Date des Premiers Symptomes de cet accès")
    row["date_premiers_symptomes"] = fmt_date(v) if v else ""

    # ── Prophylaxis ───────────────────────────────────────────────────────────
    chimio = get_cnr(all_cnr, "Chimioprophylaxie utilisée")
    if chimio and norm(chimio) not in ("non", "nsp"):
        row["chimioprophylaxie"] = chimio
    else:
        for kw, val in config.get("chimio_keywords", {}).items():
            if kw in norm(all_text):
                row["chimioprophylaxie"] = val
                break

    # ── Treatment ─────────────────────────────────────────────────────────────
    treat = get_cnr(all_cnr, "Traitement anti-palustre de 1ère intention")
    if not treat:
        found = []
        for kw, val in config.get("treatment_keywords", {}).items():
            if kw in norm(all_text) and val not in found:
                found.append(val)
        treat = " + ".join(found)
    row["Traitement_antipalu"] = treat
    row["Durée_traitement"]    = get_cnr(all_cnr, "Durée en jours")
    row["Hospitalisation"]     = get_cnr(all_cnr, "Hospitalisation")
    row["Durée_hospit"]        = get_cnr(all_cnr, "Nombre de jours d'hospitalisation")
    row["Perfusion"]           = detect_boolean(fcfg.get("Perfusion",{}), clinical_text)
    row["ECG"]                 = detect_boolean(fcfg.get("ECG",{}), all_text)

    # Type service (check J0 group first)
    j0_text = groups["J0"].text()
    for svc in config.get("institution",{}).get("services",[]):
        if norm(svc) in norm(j0_text):
            row["type_service"] = svc
            break

    # ── Malaria ───────────────────────────────────────────────────────────────
    if re.search(r"falciparum|P\.?\s*falciparum", all_text, re.IGNORECASE):
        row["Type_paludisme"] = "P. falciparum"
    else:
        row["Type_paludisme"] = get_cnr(all_cnr, "Espèce(s) Plasmodiale(s)")

    row["gravité_palu"] = detect_severity(
        fcfg.get("gravité_palu",{}), clinical_text, all_cnr
    )

    # ── J0 / J3 / J30 ────────────────────────────────────────────────────────
    for prefix in ["J0","J3","J30"]:
        fill_timepoint(row, prefix, groups[prefix], all_cnr, fcfg,
                       tp_dates.get(prefix,""))

    return row


# ─────────────────────────────────────────────────────────────────────────────
# Step 4 – Write outputs
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(rows: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️  openpyxl not found: pip install openpyxl --break-system-packages")
        return

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Données patients"
        hdr  = PatternFill("solid", fgColor="FFD966")
        col  = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="AAAAAA")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        sections_def = [
            ("Profil épidémiologique", 1, 21),
            ("J0 – Admission",        22, 62),
            ("J3 – Contrôle",         63, 103),
            ("J30 – Suivi",           104, 144),
            ("Traitement & issue",    145, len(COLUMNS)),
        ]
        for title, s, e in sections_def:
            e = min(e, len(COLUMNS))
            if s > len(COLUMNS): break
            ws.merge_cells(start_row=1,start_column=s,end_row=1,end_column=e)
            c = ws.cell(row=1,column=s)
            c.value=title; c.font=Font(bold=True,size=10); c.fill=hdr
            c.alignment=Alignment(horizontal="center",vertical="center")

        for ci, cname in enumerate(COLUMNS,1):
            c=ws.cell(row=2,column=ci)
            c.value=cname; c.font=Font(bold=True,size=9); c.fill=col
            c.border=brd; c.alignment=Alignment(horizontal="center",wrap_text=True)

        ws.freeze_panes="A3"
        ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=42
        for ci in range(1,len(COLUMNS)+1):
            ws.column_dimensions[get_column_letter(ci)].width=14
        ws.column_dimensions["A"].width=18

    for row_data in rows:
        pid=row_data["ID_Patient"]; target=None
        for r in range(3,ws.max_row+1):
            if ws.cell(r,1).value==pid: target=r; break
        if target is None: target=max(ws.max_row+1,3)
        for ci,cname in enumerate(COLUMNS,1):
            c=ws.cell(row=target,column=ci)
            c.value=row_data.get(cname,"")
            c.alignment=Alignment(vertical="top",wrap_text=False)

    wb.save(path)
    print(f"  ✅ Excel: {path}")


def write_sqlite(rows: list[dict], path: Path) -> None:
    conn = sqlite3.connect(path)
    cur  = conn.cursor()
    cols_sql = ", ".join(f'"{c}" TEXT' for c in COLUMNS)
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS patients (
            {cols_sql},
            extraction_date TEXT,
            PRIMARY KEY ("ID_Patient")
        )
    """)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for rd in rows:
        col_names    = ", ".join(f'"{c}"' for c in COLUMNS) + ', "extraction_date"'
        placeholders = ", ".join("?" for _ in range(len(COLUMNS)+1))
        values       = [rd.get(c,"") for c in COLUMNS] + [ts]
        cur.execute(f"INSERT OR REPLACE INTO patients ({col_names}) VALUES ({placeholders})",values)
    conn.commit(); conn.close()
    print(f"  ✅ SQLite: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def process_patient(anon_file: Path, config: dict) -> Optional[dict]:
    m = re.search(r"patient_(\d+)_anonymized", anon_file.name, re.IGNORECASE)
    if not m:
        print(f"⚠️  Cannot parse patient ID from: {anon_file.name}")
        return None

    patient_id = f"PATIENT_{m.group(1).zfill(3)}"
    print(f"\n{'='*60}")
    print(f"  Patient: {patient_id}  ({anon_file.name})")

    sections = parse_anonymized_file(anon_file, config)
    for sec in sections:
        dates = sorted({p.date for p in sec.pages if p.date})
        print(f"  {sec.doc_id} [{sec.doc_type}] pages={[p.num for p in sec.pages]} "
              f"dates={dates} cnr={'yes' if sec.cnr_fields else 'no'}")

    j0_merge = config.get("j0_merge_days", 3)
    groups, tp_dates = assign_timepoints(sections, j0_merge)

    print(f"\n  Timepoint assignment (merge window: ≤{j0_merge} days → J0):")
    for tp in ["J0","J3","J30"]:
        g    = groups[tp]
        docs = sorted({s.doc_id for s,_ in g.pages})
        print(f"    {tp} ({tp_dates.get(tp,'—')}): {docs or '(empty)'}")

    row = extract_row(patient_id, sections, groups, tp_dates, config)

    print(f"\n  Key extracted values:")
    show = ["Sexe","Age","Lieu_séjour","chimioprophylaxie","Type_paludisme",
            "gravité_palu","Traitement_antipalu","Hospitalisation","Durée_hospit",
            "Date_PEC_J0","Poids_J0","Température_J0","Fréquence_Cardiaque_J0",
            "hémoglobine_J0","plaquettes_J0","globules_blancs_J0",
            "ASAT_J0","ALAT_J0","CRP_J0","Parasitémie_J0",
            "Frottis_sanguin_J0","HRP2_J0",
            "Date_PEC_J3","Température_J3","Frottis_sanguin_J3",
            "Date_PEC_J30","PDV_J30","Frottis_sanguin_J30"]
    for col in show:
        val = row.get(col,"")
        if val:
            print(f"    {col}: {val}")

    return row


def main():
    parser = argparse.ArgumentParser(
        description="System B variable extraction → Excel + SQLite"
    )
    grp = parser.add_mutually_exclusive_group(required=True)
    grp.add_argument("--patient",  help="Path to patient_NNN_anonymized.txt")
    grp.add_argument("--all",      action="store_true",
                     help="Process all files in final_anonymized/")
    parser.add_argument("--config",     default=str(CONFIG_PATH))
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR))
    args = parser.parse_args()

    config  = load_config(Path(args.config))
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.patient:
        files = [Path(args.patient)]
    else:
        files = sorted(FINAL_ANON.glob("patient_*_anonymized.txt"))
        if not files:
            print(f"No files found in {FINAL_ANON}")
            sys.exit(1)
        print(f"Found {len(files)} patient file(s)")

    rows = []
    for f in files:
        row = process_patient(f, config)
        if row:
            rows.append(row)

    if rows:
        write_excel(rows, out_dir/"research_table.xlsx")
        write_sqlite(rows, out_dir/"research_database.db")
        print(f"\n{'='*60}")
        print(f"✅ Done: {len(rows)} patient(s) processed")
    else:
        print("No rows extracted."); sys.exit(1)


if __name__ == "__main__":
    main()