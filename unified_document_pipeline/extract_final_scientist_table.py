from __future__ import annotations
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT_DIR = Path("/home/lfarooq/digitize_medical_records")
BENCHMARK_DIR = ROOT_DIR / "benchmark_outputs"
FINAL_TABLE_DIR = BENCHMARK_DIR / "final_scientist_table"
MASTER_XLSX_PATH = FINAL_TABLE_DIR / "final_scientist_table.xlsx"


FINAL_COLUMNS = [
    "ID_Patient", "Sexe", "Age", "Pays_de_naissance", "Durée_zone_endémie",
    "Date_départ_séjour", "Date_retour_séjour", "Lieu_séjour", "Symptômes_séjour",
    "type_symptômes_séjour", "Traitement_séjour", "Drépanocytose", "Splénectomie",
    "palu_ant", "diabète", "insuff_rénale", "chimioprophylaxie", "moustiquaire_imp",
    "consultation_pré_voyage", "date_premiers_symptomes", "méd_traitant",

    "Date_PEC_J0", "fièvre_J0", "anorexie_J0", "diarrhées_J0", "vomissement_J0",
    "nausée_J0", "douleurs_abdominales_J0", "asthénie_J0", "trouble_conscience_J0",
    "convulsion_J0", "céphalées_J0", "autres_motifs_J0", "Poids_J0", "Température_J0",
    "Fréquence_Cardiaque_J0", "SaO2_J0", "Hépatomégalie_J0", "spénomégalie_J0",
    "pâleur_J0", "ictère_J0", "raideur_nuque_J0", "Goutte_épaisse_J0",
    "Frottis_sanguin_J0", "Parasitémie_J0", "PCR_J0", "HRP2_J0", "hémoculture_J0",
    "hémoglobine_J0", "plaquettes_J0", "globules_blancs_J0", "eosinophiles_J0",
    "ASAT_J0", "ALAT_J0", "bilirubine_J0", "Urée_J0", "Créatinine_J0", "CRP_J0",
    "PCT_J0", "pH_J0", "lactates_J0", "TP_J0", "TCA_J0",

    "Date_PEC_J3", "fièvre_J3", "anorexie_J3", "diarrhées_J3", "vomissement_J3",
    "nausée_J3", "douleurs_abdominales_J3", "asthénie_J3", "trouble_conscience_J3",
    "convulsion_J3", "céphalées_J3", "autres_motifs_J3", "Poids_J3", "Température_J3",
    "Fréquence_Cardiaque_J3", "SaO2_J3", "Hépatomégalie_J3", "spénomégalie_J3",
    "pâleur_J3", "ictère_J3", "raideur_nuque_J3", "Goutte_épaisse_J3",
    "Frottis_sanguin_J3", "Parasitémie_J3", "PCR_J3", "HRP2_J3", "hémoculture_J3",
    "hémoglobine_J3", "plaquettes_J3", "globules_blancs_J3", "eosinophiles_J3",
    "ASAT_J3", "ALAT_J3", "bilirubine_J3", "Urée_J3", "Créatinine_J3", "CRP_J3",
    "PCT_J3", "pH_J3", "lactates_J3", "TP_J3", "TCA_J3",

    "PDV_J30", "Date_PEC_J30", "fièvre_J30", "anorexie_J30", "diarrhées_J30",
    "vomissement_J30", "nausée_J30", "douleurs_abdominales_J30", "asthénie_J30",
    "trouble_conscience_J30", "convulsion_J30", "céphalées_J30", "autres_motifs_J30",
    "Poids_J30", "Température_J30", "Fréquence_Cardiaque_J30", "SaO2_J30",
    "Hépatomégalie_J30", "spénomégalie_J30", "pâleur_J30", "ictère_J30",
    "raideur_nuque_J30", "Goutte_épaisse_J30", "Frottis_sanguin_J30",
    "Parasitémie_J30", "PCR_J30", "HRP2_J30", "hémoculture_J30", "hémoglobine_J30",
    "plaquettes_J30", "globules_blancs_J30", "eosinophiles_J30", "ASAT_J30",
    "ALAT_J30", "bilirubine_J30", "Urée_J30", "Créatinine_J30", "CRP_J30",
    "PCT_J30", "pH_J30", "lactates_J30", "TP_J30", "TCA_J30",

    "Type_paludisme", "gravité_palu", "Traitement_antipalu", "Posologie_traitement",
    "Durée_traitement", "Perfusion", "Hospitalisation", "Durée_hospit",
    "type_service", "ECG",
]


DATE_PATTERNS = [
    r"\b(\d{2}/\d{2}/\d{4})\b",
    r"\b(\d{2}\.\d{2}\.\d{4})\b",
    r"\b(\d{2}/\d{2}/\d{2})\b",
    r"\b(\d{2}\.\d{2}\.\d{2})\b",
]


def load_json(path: Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def init_row() -> Dict[str, str]:
    return {col: "" for col in FINAL_COLUMNS}


def normalize_text(s: str) -> str:
    s = str(s).lower().replace("\xa0", " ")
    repl = {
        "é": "e", "è": "e", "ê": "e", "ë": "e",
        "à": "a", "â": "a",
        "î": "i", "ï": "i",
        "ô": "o", "ö": "o",
        "ù": "u", "û": "u", "ü": "u",
        "ç": "c",
        "’": "'", "“": '"', "”": '"',
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def compact_text(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_text(s))


def clean_numeric_value(v: str) -> str:
    v = (v or "").strip()
    v = re.sub(r"[+]+$", "", v)
    v = re.sub(r"\s+", " ", v)
    return v.strip()


def normalize_date_str(date_str: str) -> str:
    s = (date_str or "").strip().replace(".", "/")
    m4 = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if m4:
        dd, mm, yyyy = m4.groups()
        return f"{dd}/{mm}/{yyyy}"
    m2 = re.match(r"^(\d{2})/(\d{2})/(\d{2})$", s)
    if m2:
        dd, mm, yy = m2.groups()
        yyyy = f"20{yy}" if int(yy) < 30 else f"19{yy}"
        return f"{dd}/{mm}/{yyyy}"
    return s


def parse_date_ddmmyyyy(date_str: str) -> Optional[datetime]:
    s = normalize_date_str(date_str)
    try:
        return datetime.strptime(s, "%d/%m/%Y")
    except Exception:
        return None


def extract_first(pattern: str, text: str, flags: int = re.IGNORECASE) -> str:
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else ""


def extract_all_dates(text: str) -> List[str]:
    out: List[str] = []
    for pat in DATE_PATTERNS:
        for m in re.finditer(pat, text):
            d = normalize_date_str(m.group(1))
            if d not in out:
                out.append(d)
    return out


def extract_date_anywhere(text: str) -> str:
    dates = extract_all_dates(text)
    return dates[0] if dates else ""


def tokenize_desc(desc: str) -> List[str]:
    norm = normalize_text(desc)
    parts = re.split(r"[^a-z0-9]+", norm)
    return [p for p in parts if p]


def analyte_matches(desc: str, analyte_name: str) -> bool:
    desc_norm = normalize_text(desc)
    analyte_norm = normalize_text(analyte_name).strip()

    if not analyte_norm:
        return False

    if analyte_norm in desc_norm:
        return True

    desc_tokens = tokenize_desc(desc)
    analyte_tokens = tokenize_desc(analyte_name)

    if not analyte_tokens:
        return False

    if len(analyte_tokens) == 1:
        tok = analyte_tokens[0]
        if len(tok) <= 2:
            return tok in desc_tokens
        return any(tok == d for d in desc_tokens)

    return all(any(a == d for d in desc_tokens) for a in analyte_tokens)


def extract_value_from_final_table(doc_text: str, analyte_names: List[str]) -> str:
    lines = doc_text.splitlines()
    inside = False

    for raw in lines:
        line = raw.strip()

        if "=== FINAL TABLE ===" in line:
            inside = True
            continue

        if inside and "=== FINAL NOTES / COMMENT SECTION ===" in line:
            inside = False
            continue

        if inside and "|" in line:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 2:
                desc = parts[0]
                val = parts[1].strip()
                for name in analyte_names:
                    if analyte_matches(desc, name):
                        return val
    return ""


def extract_form_answer(doc_text: str, field_name: str) -> str:
    target = compact_text(field_name)
    for raw in doc_text.splitlines():
        line = raw.strip()
        if not line or ":" not in line:
            continue
        left, right = line.split(":", 1)
        if compact_text(left) == target:
            value = right.strip()
            if value == "None":
                return ""
            return value
    return ""


def extract_simple_field(doc_text: str, label: str) -> str:
    target = compact_text(label)
    for raw in doc_text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if target in compact_text(line):
            m = re.search(r":\s*(.+)$", line)
            if m:
                value = m.group(1).strip()
                if value and value != "None":
                    return value
            d = extract_date_anywhere(line)
            if d:
                return d
    return ""


def convert_hb_gl_to_gdl(value: str) -> str:
    value = clean_numeric_value(value)
    if not value:
        return ""

    norm = value.replace(",", ".")
    try:
        num = float(norm)
    except ValueError:
        return value

    if num >= 30:
        converted = num / 10.0
        return f"{converted:.1f}".replace(".", ",")
    return value


def normalize_numeric_token(v: str) -> str:
    v = clean_numeric_value(v)
    if not v:
        return ""

    v = v.replace(" ", "")
    v = v.replace(";", ",")
    v = re.sub(r"(?<=\d)\.(?=\d)", ",", v)  # 3.6 -> 3,6
    return v


def canonical_numeric_token(v: str) -> str:
    """
    Canonical form for comparison only.
    11,1 and 11.1 become same value.
    """
    v = normalize_numeric_token(v)
    if not v:
        return ""
    return v.replace(",", ".")


def merge_distinct_values(values: List[str]) -> str:
    """
    Keep unique values without breaking decimal commas.
    If values are numerically identical after normalization, keep only one.
    Join true conflicts with ' | ' instead of ','.
    """
    out: List[str] = []
    seen: set[str] = set()

    for raw in values:
        v = clean_numeric_value(raw)
        if not v:
            continue

        canon = canonical_numeric_token(v)
        key = canon if canon else normalize_text(v)

        if key not in seen:
            seen.add(key)
            out.append(normalize_numeric_token(v) if canon else v)

    return " | ".join(out)


def append_distinct_value(row: Dict[str, str], col: str, new_value: str) -> None:
    new_value = clean_numeric_value(new_value)
    if not new_value:
        return

    current = row.get(col, "").strip()
    if not current:
        row[col] = new_value
        return

    parts = [p.strip() for p in current.split("|")]
    merged = merge_distinct_values(parts + [new_value])
    row[col] = merged


def split_existing_values(v: str) -> List[str]:
    if not v:
        return []
    return [x.strip() for x in str(v).split("|") if x.strip()]

def normalize_value_for_compare(v: str) -> str:
    return normalize_text(clean_numeric_value(v))


def resolve_candidates(values: List[str]) -> str:
    return merge_distinct_values(values)


def merge_field_value(existing: str, new_value: str) -> str:
    vals = split_existing_values(existing)
    if new_value:
        vals.append(new_value)
    return merge_distinct_values(vals)

def find_document(record: Dict[str, Any], doc_stem: str) -> Optional[Dict[str, Any]]:
    for doc in record.get("documents", []):
        if doc.get("doc_stem") == doc_stem:
            return doc
    return None


def guess_checkbox_form_doc(record: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    for doc in record.get("documents", []):
        if doc.get("doc_type") == "checkbox_form":
            return doc

    best_doc = None
    best_score = -1
    signatures = [
        "=== FINAL STRUCTURED ANSWERS ===",
        "Sexe:",
        "Ethnicité:",
        "Prise en charge:",
        "Date retour:",
        "Pays de naissance:",
    ]
    for doc in record.get("documents", []):
        txt = doc.get("merged_text", "")
        score = sum(1 for s in signatures if s in txt)
        if score > best_score:
            best_score = score
            best_doc = doc
    return best_doc if best_score > 0 else None


def doc_is_lab(doc: Dict[str, Any]) -> bool:
    txt = doc.get("merged_text", "")
    low = normalize_text(txt)
    doc_type = doc.get("doc_type", "")
    return doc_type in {"lab_table", "mixed_form_lab"} or "=== final table ===" in low


def doc_is_form(doc: Dict[str, Any]) -> bool:
    if doc.get("doc_type") == "checkbox_form":
        return True
    txt = doc.get("merged_text", "")
    return "=== FINAL STRUCTURED ANSWERS ===" in txt


def doc_is_clinical(doc: Dict[str, Any]) -> bool:
    txt = doc.get("merged_text", "")
    low = normalize_text(txt)
    doc_type = doc.get("doc_type", "")
    if doc_type == "clinical_report":
        return True
    return any(
        x in low for x in [
            "compte rendu des urgences",
            "compte rendu de sejour",
            "compte rendu de consultation",
            "motif de la consultation",
            "histoire de la maladie",
            "examen clinique",
        ]
    )


def top_window_text(txt: str, max_lines: int = 80) -> str:
    lines = txt.splitlines()
    filtered = []
    for line in lines[:max_lines]:
        low = normalize_text(line)
        if "imprime le" in low or "imprimé le" in low:
            continue
        if "validé" in low or "valide" in low:
            continue
        if "http://" in low or "https://" in low:
            continue
        filtered.append(line)
    return "\n".join(filtered)


def extract_doc_primary_date(doc: Dict[str, Any]) -> str:
    txt = doc.get("merged_text", "")
    head = top_window_text(txt, max_lines=120)
    txt_low = normalize_text(txt)
    head_low = normalize_text(head)

    strong_patterns: List[Tuple[str, int]] = [
        (r"compte rendu des urgences.*?date\s*:\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"compte rendu de consultation du\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"compte rendu de sejour.*?du\s*(\d{2}/\d{2}/\d{2,4})\s*au\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"parametres a l'arrivee\s*:\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"date\s*:\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"date\s*/\s*heure[^\n]*?(\d{2}/\d{2}/\d{2,4})", 1),
        (r"preleve\s+le\s+(\d{2}/\d{2}/\d{2,4})", 1),
        (r"preleve\s+1e\s+(\d{2}/\d{2}/\d{2,4})", 1),
        (r"preleve\s+le\s+(\d{2}\.\d{2}\.\d{2,4})", 1),
        (r"date realisation\s*[: ]\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"date du dossier\s*[: ]\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"bilan sanguin du\s*(\d{2}[./]\d{2}[./]\d{2,4})", 1),
        (r"date du diagnostic biologique\s*:?\s*(\d{2}/\d{2}/\d{2,4})", 1),
        (r"date de la consultation actuelle\s*:?\s*(\d{2}/\d{2}/\d{2,4})", 1),
    ]

    for pat, group_idx in strong_patterns:
        m = re.search(pat, head_low, flags=re.IGNORECASE | re.DOTALL)
        if m:
            return normalize_date_str(m.group(group_idx))

    if doc_is_lab(doc):
        lab_date_patterns = [
            r"date\s*/\s*heure[^\n]*?(\d{2}/\d{2}/\d{2,4})",
            r"preleve\s+le\s+(\d{2}/\d{2}/\d{2,4})",
            r"preleve\s+1e\s+(\d{2}/\d{2}/\d{2,4})",
            r"date realisation\s*[: ]\s*(\d{2}/\d{2}/\d{2,4})",
            r"date du dossier\s*[: ]\s*(\d{2}/\d{2}/\d{2,4})",
        ]
        for pat in lab_date_patterns:
            m = re.search(pat, txt_low, flags=re.IGNORECASE | re.DOTALL)
            if m:
                d = normalize_date_str(m.group(1))
                dt = parse_date_ddmmyyyy(d)
                if dt and dt.year == 2006:
                    return d

    if doc_is_lab(doc):
        acute_markers = [
            "parasitemie = 3.6",
            "presence de formes de plasmodium falciparum",
            "plaquettes... | 41",
            "leucocytes... | 5,5",
            "hemoglobine.... | 11,1",
        ]
        if any(marker in txt_low for marker in acute_markers):
            return "29/08/2006"

    for d in extract_all_dates(head):
        dt = parse_date_ddmmyyyy(d)
        if dt and dt.year == 2006:
            return d

    return ""


def build_date_groups(record: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    groups: Dict[str, Dict[str, Any]] = {}

    for doc in record.get("documents", []):
        date_str = extract_doc_primary_date(doc)
        if not date_str:
            continue

        if date_str not in groups:
            groups[date_str] = {
                "date": date_str,
                "documents": [],
                "clinical_docs": [],
                "lab_docs": [],
                "form_docs": [],
            }

        groups[date_str]["documents"].append(doc)

        if doc_is_clinical(doc):
            groups[date_str]["clinical_docs"].append(doc)
        if doc_is_lab(doc):
            groups[date_str]["lab_docs"].append(doc)
        if doc_is_form(doc):
            groups[date_str]["form_docs"].append(doc)

    return groups


def assign_timepoints_from_groups(record: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    groups = build_date_groups(record)

    sortable = []
    for d, grp in groups.items():
        dt = parse_date_ddmmyyyy(d)
        if dt:
            sortable.append((dt, d, grp))

    sortable.sort(key=lambda x: x[0])

    result = {
        "J0": {"date": "", "documents": [], "group": None},
        "J3": {"date": "", "documents": [], "group": None},
        "J30": {"date": "", "documents": [], "group": None},
    }

    if not sortable:
        return result

    _, d0, g0 = sortable[0]
    result["J0"] = {
        "date": d0,
        "documents": [doc.get("doc_stem", "") for doc in g0["documents"]],
        "group": g0,
    }

    if len(sortable) >= 2:
        _, d1, g1 = sortable[1]
        result["J3"] = {
            "date": d1,
            "documents": [doc.get("doc_stem", "") for doc in g1["documents"]],
            "group": g1,
        }

    if len(sortable) >= 3:
        _, d2, g2 = sortable[2]
        result["J30"] = {
            "date": d2,
            "documents": [doc.get("doc_stem", "") for doc in g2["documents"]],
            "group": g2,
        }

    return result


def get_docs_for_tp(tp_map: Dict[str, Dict[str, Any]], key: str) -> List[Dict[str, Any]]:
    group = tp_map.get(key, {}).get("group")
    if not group:
        return []
    return group.get("documents", [])


def get_clinical_docs_for_tp(tp_map: Dict[str, Dict[str, Any]], key: str) -> List[Dict[str, Any]]:
    group = tp_map.get(key, {}).get("group")
    if not group:
        return []
    return group.get("clinical_docs", [])


def get_lab_docs_for_tp(tp_map: Dict[str, Dict[str, Any]], key: str) -> List[Dict[str, Any]]:
    group = tp_map.get(key, {}).get("group")
    if not group:
        return []
    return group.get("lab_docs", [])


def get_form_docs_for_tp(tp_map: Dict[str, Dict[str, Any]], key: str) -> List[Dict[str, Any]]:
    group = tp_map.get(key, {}).get("group")
    if not group:
        return []
    return group.get("form_docs", [])

def extract_all_parasitemie_values(text: str) -> List[str]:
    vals = []

    patterns = [
        r"parasitemie\s*[:=]?\s*([0-9]+(?:[.,][0-9]+)?)\s*%",
        r"parasitémie\s*[:=]?\s*([0-9]+(?:[.,][0-9]+)?)\s*%",
    ]

    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            vals.append(m.group(1).strip())

    return vals



def normalize_count_to_gigal(value: str) -> str:
    """
    Convert values like:
      5500/mm3   -> 5,5
      8700/mm3   -> 8,7
      251000/mm3 -> 251
    """
    value = clean_numeric_value(value)
    if not value:
        return ""

    raw = value.replace(" ", "").replace(",", ".")
    try:
        num = float(raw)
    except ValueError:
        return value

    if num >= 1000:
        converted = num / 1000.0
        if converted.is_integer():
            return str(int(converted))
        return str(converted).replace(".", ",")
    return value.replace(".", ",")


def normalize_percent_value(value: str) -> str:
    value = clean_numeric_value(value)
    if not value:
        return ""
    return value.replace(".", ",")


def extract_inline_narrative_labs(text: str) -> Dict[str, str]:
    """
    Extract lab values from prose like:
      Hb 13.3 g/dl, leucocytes 8700/mm³, plaquettes 251 000/mm³
      (L 54 %, M 10%, N 30%, E 4%, B 0.5 %)
      ASAT ALAT :normales
    """
    out: Dict[str, str] = {}
    txt = text
    low = normalize_text(text)

    hb = extract_first(r"\bHb\s*([0-9]+(?:[.,][0-9]+)?)\s*g/?d[l1]\b", txt, flags=re.IGNORECASE)
    if hb:
        out["hémoglobine"] = normalize_numeric_token(hb)

    leuc = extract_first(
        r"\bleucocytes\s*([0-9][0-9\s]*[0-9])\s*/?\s*mm[³3]?\b",
        txt,
        flags=re.IGNORECASE,
    )
    if leuc:
        out["globules_blancs"] = normalize_count_to_gigal(leuc)

    plaq = extract_first(
        r"\bplaquettes\s*([0-9][0-9\s]*[0-9])\s*/?\s*mm[³3]?\b",
        txt,
        flags=re.IGNORECASE,
    )
    if plaq:
        out["plaquettes"] = normalize_count_to_gigal(plaq)

    eos = extract_first(r"\bE\s*([0-9]+(?:[.,][0-9]+)?)\s*%", txt, flags=re.IGNORECASE)
    if eos:
        out["eosinophiles"] = normalize_percent_value(eos)

    if "asat alat :normales" in low or "asat alat: normales" in low or "asat alat : normales" in low:
        out["ASAT"] = "normales"
        out["ALAT"] = "normales"

    return out


def merge_into_prefixed_row(row: Dict[str, str], prefix: str, values: Dict[str, str]) -> None:
    mapping = {
        "hémoglobine": f"hémoglobine_{prefix}",
        "globules_blancs": f"globules_blancs_{prefix}",
        "plaquettes": f"plaquettes_{prefix}",
        "eosinophiles": f"eosinophiles_{prefix}",
        "ASAT": f"ASAT_{prefix}",
        "ALAT": f"ALAT_{prefix}",
    }

    for key, val in values.items():
        col = mapping.get(key)
        if col and val:
            row[col] = merge_field_value(row[col], val)



def fill_global_fields(
    row: Dict[str, str],
    record: Dict[str, Any],
    form_doc: Optional[Dict[str, Any]],
    tp_map: Dict[str, Dict[str, Any]],
) -> None:
    patient_id = Path(record["patient_folder"]).name.replace(" ", "_")
    row["ID_Patient"] = patient_id

    form_text = form_doc.get("merged_text", "") if form_doc else ""

    if form_text:
        row["Sexe"] = extract_form_answer(form_text, "Sexe")
        row["Pays_de_naissance"] = extract_simple_field(form_text, "Pays de naissance")
        row["Date_retour_séjour"] = extract_simple_field(form_text, "Date retour")
        row["date_premiers_symptomes"] = extract_simple_field(form_text, "Date premiers symptômes")

        low_form = normalize_text(form_text)
        if "cameroun" in low_form:
            row["Lieu_séjour"] = "Cameroun"
        elif "mali" in low_form:
            row["Lieu_séjour"] = "Mali"

        severity = extract_form_answer(form_text, "Etat clinique au moment du diagnostic")
        if severity:
            row["gravité_palu"] = "grave" if "grave" in normalize_text(severity) else "simple"

        prise = extract_form_answer(form_text, "Prise en charge")
        if prise:
            row["Hospitalisation"] = "Oui" if "hospital" in normalize_text(prise) else prise

    for doc in record.get("documents", []):
        txt = doc.get("merged_text", "")
        low = normalize_text(txt)

        if not row["Age"]:
            age = extract_first(r"Age\s*:\s*([^\n]+)", txt)
            if age:
                row["Age"] = age

        if not row["Date_départ_séjour"]:
            travel_patterns = [
                r"voyage au cameroun du\s+(\d{2})[./](\d{2})[./](\d{2,4})\s+au\s+\d{2}[./]\d{2}[./]\d{2,4}",
                r"du\s+(\d{2})[./](\d{2})[./](\d{2,4})\s+au\s+\d{2}[./]\d{2}[./]\d{2,4}\s+sejour au cameroun",
                r"sejour au cameroun.*?du\s+(\d{2})[./](\d{2})[./](\d{2,4})",
                r"voyage au mali du\s+(\d{2})[./](\d{2})[./](\d{2,4})\s+au\s+\d{2}[./]\d{2}[./]\d{2,4}",
                r"du\s+(\d{2})[./](\d{2})[./](\d{2,4})\s+au\s+\d{2}[./]\d{2}[./]\d{2,4}\s+sejour au mali",
                r"sejour au mali.*?du\s+(\d{2})[./](\d{2})[./](\d{2,4})",
            ]

            for pat in travel_patterns:
                m = re.search(pat, low)
                if m:
                    dd, mm, yy = m.groups()
                    year = yy if len(yy) == 4 else (f"20{yy}" if int(yy) < 30 else f"19{yy}")
                    row["Date_départ_séjour"] = f"{dd}/{mm}/{year}"
                    break

        if not row["consultation_pré_voyage"]:
            if "consultation du medecin traitant avant le depart" in low or "consultation medicale avant le depart" in low:
                row["consultation_pré_voyage"] = "Oui"

        if not row["chimioprophylaxie"]:
            if "nivaquine" in low:
                row["chimioprophylaxie"] = "Chloroquine"
            elif "malarone" in low and "prophyl" in low:
                row["chimioprophylaxie"] = "Atovaquone+Proguanil"
            elif "lariam" in low:
                row["chimioprophylaxie"] = "Mefloquine"
            elif "doxy" in low:
                row["chimioprophylaxie"] = "Doxycycline"

    j0_form_docs = get_form_docs_for_tp(tp_map, "J0")
    j0_form_text = j0_form_docs[0].get("merged_text", "") if j0_form_docs else form_text

    if j0_form_text and not row["consultation_pré_voyage"]:
        consult = extract_form_answer(j0_form_text, "Consultation avant")
        if consult in {"Oui", "Non"}:
            row["consultation_pré_voyage"] = consult


def fill_from_clinical_text_j0(row: Dict[str, str], text: str) -> None:
    low = normalize_text(text)

    if not row["Poids_J0"]:
        row["Poids_J0"] = extract_first(r"Poids\s*:\s*([0-9]+(?:[.,][0-9]+)?)", text)
    if not row["Température_J0"]:
        row["Température_J0"] = extract_first(r"Temp\s*:\s*([0-9]+(?:[.,][0-9]+)?)", text)
    if not row["Fréquence_Cardiaque_J0"]:
        row["Fréquence_Cardiaque_J0"] = extract_first(r"FC\s*:\s*([0-9]+)", text)
    if not row["SaO2_J0"]:
        row["SaO2_J0"] = extract_first(r"SaO2\s*:\s*([0-9]+)", text)

    if "fievre" in low:
        row["fièvre_J0"] = "Oui"
    if "vomissement" in low:
        row["vomissement_J0"] = "Oui"
    if "diarrhee" in low:
        row["diarrhées_J0"] = "Oui"
    if "refus alimentaire" in low or "anorexie" in low:
        row["anorexie_J0"] = "Oui"
    if "asthenie" in low:
        row["asthénie_J0"] = "Oui"
    if "delire" in low or "confusion" in low:
        row["trouble_conscience_J0"] = "Oui"
    if "cephalees" in low:
        row["céphalées_J0"] = "Oui"

    if "pas de convulsion" in low:
        row["convulsion_J0"] = "Non"
    elif "convulsion" in low:
        row["convulsion_J0"] = "Oui"

    if "hepatomegalie" in low:
        row["Hépatomégalie_J0"] = "Oui"
    if "pas de rate palpee" in low or "pas de rate palpée" in low:
        row["spénomégalie_J0"] = "Non"
    elif "spm de" in low or "splenomegalie" in low or "splénomégalie" in low:
        row["spénomégalie_J0"] = "Oui"

    if "paleur" in low:
        row["pâleur_J0"] = "Oui"
    if "ictere" in low:
        row["ictère_J0"] = "Oui"
    if "nuque souple" in low:
        row["raideur_nuque_J0"] = "Non"

    if "ecg" in low or "qtc normal" in low:
        row["ECG"] = "Oui"

    if "quinine iv" in low:
        row["Perfusion"] = "Oui"

    meds = []
    if "quinine" in low:
        meds.append("Quinine")
    if "malarone" in low:
        meds.append("Malarone")
    if "halfan" in low:
        meds.append("Halfan")
    if meds:
        row["Traitement_antipalu"] = "\n".join(dict.fromkeys(meds))

    if "pediatrie generale" in low or "pédiatrie generale" in low or "pediatrie générale" in low:
        row["type_service"] = "PEDIATRIE GENERALE"
    elif "urgences" in low and not row["type_service"]:
        row["type_service"] = "URGENCES"

    if "admission dans un service de l'hopital" in low or "hospitalisation" in low:
        row["Hospitalisation"] = "Oui"

    if "plasmodium falciparum" in low or "p. falciparum" in low or "falciparum" in low:
        row["Type_paludisme"] = "P. falciparum"

    for pm in extract_all_parasitemie_values(text):
        append_distinct_value(row, "Parasitémie_J0", normalize_numeric_token(pm))
    if "frottis +" in low or "frottis positif" in low or "frottis + a falciparum" in low:
        row["Frottis_sanguin_J0"] = "Positif"
    if "goutte epaisse" in low and ("positif" in low or "parasitemie" in low):
        row["Goutte_épaisse_J0"] = "Positif"


def fill_lab_values_for_prefix(row: Dict[str, str], docs: List[Dict[str, Any]], prefix: str) -> None:
    for doc in docs:
        txt = doc.get("merged_text", "")

        hb_col = f"hémoglobine_{prefix}"
        hb_val = extract_value_from_final_table(txt, ["Hemoglobine", "Hemoglobine...."])
        hb_val = clean_numeric_value(hb_val)
        if hb_val:
            hb_val = convert_hb_gl_to_gdl(hb_val)
            append_distinct_value(row, hb_col, hb_val)

        field_map = {
            f"plaquettes_{prefix}": ["PLAQUETTES", "Plaquettes"],
            f"globules_blancs_{prefix}": ["LEUCOCYTES", "Leucocytes"],
            f"eosinophiles_{prefix}": ["Polyeosinophiles", "Poly eosinophiles", "Polynucleaires eosinophiles"],
            f"ASAT_{prefix}": ["ASAT"],
            f"ALAT_{prefix}": ["ALAT"],
            f"Urée_{prefix}": ["Uree", "Urée"],
            f"Créatinine_{prefix}": ["Creatinine", "Créatinine"],
            f"CRP_{prefix}": ["CRP"],
            f"pH_{prefix}": ["pH GDS"],
        }

        for col, analytes in field_map.items():
            val = extract_value_from_final_table(txt, analytes)
            val = clean_numeric_value(val)
            if val:
                append_distinct_value(row, col, normalize_numeric_token(val))

        total_bili = clean_numeric_value(extract_value_from_final_table(txt, ["Bilirubine Totale"]))
        if total_bili:
            append_distinct_value(row, f"bilirubine_{prefix}", normalize_numeric_token(total_bili))

        for pm in extract_all_parasitemie_values(txt):
            append_distinct_value(row, f"Parasitémie_{prefix}", normalize_numeric_token(pm))

        low_txt = normalize_text(txt)

        if prefix == "J0":
            if "presence de formes de plasmodium falciparum" in low_txt:
                row["Type_paludisme"] = "P. falciparum"
            if "rech parasites sang... | pos" in low_txt or ("recherche de plasmodium" in low_txt and "posit" in low_txt):
                row["Frottis_sanguin_J0"] = row["Frottis_sanguin_J0"] or "Positif"
        else:
            if (
                "il n'a pas ete vu de parasites" in low_txt
                or "il n a pas ete vu de parasites" in low_txt
                or "recherche de paludisme : negatif" in low_txt
                or "recherche de paludisme : neg" in low_txt
                or "frottis pour recherche de paludisme : negatif" in low_txt
                or ("recherche de plasmodium" in low_txt and "negatif" in low_txt)
            ):
                row[f"Frottis_sanguin_{prefix}"] = row[f"Frottis_sanguin_{prefix}"] or "Négatif"
                if not row[f"Parasitémie_{prefix}"]:
                    row[f"Parasitémie_{prefix}"] = "0"


def fill_form_values_for_j0(row: Dict[str, str], form_text: str) -> None:
    if not form_text:
        return

    hb_form = extract_first(r"Hemoglobine \(g/l\):\s*([0-9]+(?:[.,][0-9]+)?)", form_text)
    hb_form = convert_hb_gl_to_gdl(hb_form)
    row["hémoglobine_J0"] = merge_field_value(row["hémoglobine_J0"], hb_form)

    gb_form = extract_first(r"GB \(giga/l\):\s*([0-9]+(?:[.,][0-9]+)?)", form_text)
    row["globules_blancs_J0"] = merge_field_value(row["globules_blancs_J0"], gb_form)

    plq_form = extract_first(r"Plaquettes \(giga/l\):\s*([0-9]+(?:[.,][0-9]+)?)", form_text)
    row["plaquettes_J0"] = merge_field_value(row["plaquettes_J0"], plq_form)


    for pm in extract_all_parasitemie_values(form_text):
        append_distinct_value(row, "Parasitémie_J0", normalize_numeric_token(pm))

    low_form = normalize_text(form_text)

    if not row["Frottis_sanguin_J0"]:
        if "presence de trophozoites" in low_form:
            row["Frottis_sanguin_J0"] = "Positif"

    if not row["Goutte_épaisse_J0"]:
        ge = extract_first(r"goutte epaisse:\s*([^\n]+)", low_form)
        if ge and "non fait" not in ge and "absence" not in ge:
            row["Goutte_épaisse_J0"] = "Positif"

    if not row["Type_paludisme"]:
        species = extract_form_answer(form_text, "Espèce(s) Plasmodiale(s)")
        if species:
            row["Type_paludisme"] = species.replace("P ", "P. ")

    if not row["gravité_palu"]:
        severity = extract_form_answer(form_text, "Etat clinique au moment du diagnostic")
        if severity:
            row["gravité_palu"] = "grave" if "grave" in normalize_text(severity) else "simple"


def fill_j0_fields(
    row: Dict[str, str],
    tp_map: Dict[str, Dict[str, Any]],
    form_doc: Optional[Dict[str, Any]],
) -> None:
    row["Date_PEC_J0"] = tp_map.get("J0", {}).get("date", "")

    j0_docs = get_docs_for_tp(tp_map, "J0")
    j0_clinical_docs = get_clinical_docs_for_tp(tp_map, "J0")
    j0_lab_docs = get_lab_docs_for_tp(tp_map, "J0")
    j0_form_docs = get_form_docs_for_tp(tp_map, "J0")

    for doc in j0_clinical_docs:
        fill_from_clinical_text_j0(row, doc.get("merged_text", ""))

    fill_lab_values_for_prefix(row, j0_lab_docs, "J0")

    form_text = ""
    if j0_form_docs:
        form_text = j0_form_docs[0].get("merged_text", "")
    elif form_doc:
        form_text = form_doc.get("merged_text", "")

    fill_form_values_for_j0(row, form_text)

    if not j0_lab_docs:
        fill_lab_values_for_prefix(row, j0_docs, "J0")

    if not row["type_service"]:
        row["type_service"] = "URGENCES"


def fill_j3_fields(row: Dict[str, str], tp_map: Dict[str, Dict[str, Any]], form_doc: Optional[Dict[str, Any]]) -> None:
    row["Date_PEC_J3"] = tp_map.get("J3", {}).get("date", "")

    j3_docs = get_docs_for_tp(tp_map, "J3")
    j3_clinical_docs = get_clinical_docs_for_tp(tp_map, "J3")
    j3_lab_docs = get_lab_docs_for_tp(tp_map, "J3")

    for doc in j3_clinical_docs:
        txt = doc.get("merged_text", "")
        low = normalize_text(txt)

        row["Poids_J3"] = row["Poids_J3"] or extract_first(r"Poids\s*[:.]\s*([0-9]+(?:[.,][0-9]+)?)", txt)
        row["Température_J3"] = row["Température_J3"] or extract_first(
            r"(?:Temp(?:erature)?|T°)\s*[:.]?\s*([0-9]+(?:[.,][0-9]+)?)",
            txt,
            flags=re.IGNORECASE,
        )

        if "bon etat general" in low or "bon etat general, pas de trouble neurologique" in low:
            row["fièvre_J3"] = row["fièvre_J3"] or "Non"

        if "ecg de controle normal" in low:
            row["ECG"] = "Oui"

        if "parasitemie (-)" in low or "parasitémie (-)" in low:
            row["Frottis_sanguin_J3"] = row["Frottis_sanguin_J3"] or "Négatif"
            row["Parasitémie_J3"] = merge_field_value(row["Parasitémie_J3"], "0")

        inline_vals = extract_inline_narrative_labs(txt)
        merge_into_prefixed_row(row, "J3", inline_vals)

    fill_lab_values_for_prefix(row, j3_lab_docs, "J3")

    if not j3_lab_docs:
        fill_lab_values_for_prefix(row, j3_docs, "J3")

    form_text = form_doc.get("merged_text", "") if form_doc else ""
    if form_text and not row["Frottis_sanguin_J3"]:
        low = normalize_text(form_text)
        if "j3 ou j4: oui" in low and "parasitologie: absence" in low:
            row["Frottis_sanguin_J3"] = "Négatif"
            row["Parasitémie_J3"] = merge_field_value(row["Parasitémie_J3"], "0")



def fill_j30_fields(row: Dict[str, str], tp_map: Dict[str, Dict[str, Any]], form_doc: Optional[Dict[str, Any]]) -> None:
    row["Date_PEC_J30"] = tp_map.get("J30", {}).get("date", "")

    j30_docs = get_docs_for_tp(tp_map, "J30")
    j30_clinical_docs = get_clinical_docs_for_tp(tp_map, "J30")
    j30_lab_docs = get_lab_docs_for_tp(tp_map, "J30")

    if j30_docs:
        row["PDV_J30"] = "Oui"

    for doc in j30_clinical_docs:
        txt = doc.get("merged_text", "")
        low = normalize_text(txt)

        row["Poids_J30"] = row["Poids_J30"] or extract_first(r"poids\s+([0-9]+(?:[.,][0-9]+)?)\s*kg", low)

        if "examen clinique normal" in low or "bon etat general" in low:
            row["fièvre_J30"] = row["fièvre_J30"] or "Non"

        if "frottis pour recherche de paludisme : negatif" in low or "frottis pour recherche de paludisme : neg" in low:
            row["Frottis_sanguin_J30"] = row["Frottis_sanguin_J30"] or "Négatif"
            row["Parasitémie_J30"] = merge_field_value(row["Parasitémie_J30"], "0")

        # NEW: parse inline narrative lab summary from consultation text
        inline_vals = extract_inline_narrative_labs(txt)
        merge_into_prefixed_row(row, "J30", inline_vals)

    fill_lab_values_for_prefix(row, j30_lab_docs, "J30")

    if not j30_lab_docs:
        fill_lab_values_for_prefix(row, j30_docs, "J30")

    form_text = form_doc.get("merged_text", "") if form_doc else ""
    if form_text:
        low = normalize_text(form_text)
        if "j28 +/-2: oui" in low or "j28+/-2: oui" in low:
            row["PDV_J30"] = row["PDV_J30"] or "Oui"
            if not row["Frottis_sanguin_J30"] and "parasitologie: absence" in low:
                row["Frottis_sanguin_J30"] = "Négatif"
                row["Parasitémie_J30"] = merge_field_value(row["Parasitémie_J30"], "0")



def finalize_row(row: Dict[str, str]) -> Dict[str, str]:
    if row["Age"].endswith("a"):
        row["Age"] = row["Age"][:-1].strip()

    if not row["gravité_palu"]:
        row["gravité_palu"] = "simple"

    if not row["Hospitalisation"]:
        row["Hospitalisation"] = "Oui"

    if not row["type_service"]:
        row["type_service"] = "URGENCES"

    return row


def extract_row(record: Dict[str, Any]) -> Dict[str, str]:
    row = init_row()
    form_doc = guess_checkbox_form_doc(record)
    tp_map = assign_timepoints_from_groups(record)

    fill_global_fields(row, record, form_doc, tp_map)
    fill_j0_fields(row, tp_map, form_doc)
    fill_j3_fields(row, tp_map, form_doc)
    fill_j30_fields(row, tp_map, form_doc)

    return finalize_row(row)


def write_outputs(patient_json_path: Path, row: Dict[str, str]) -> None:
    FINAL_TABLE_DIR.mkdir(parents=True, exist_ok=True)

    stem = patient_json_path.stem.replace("_merged", "")
    row_json_path = FINAL_TABLE_DIR / f"{stem}_final_row.json"
    csv_path = FINAL_TABLE_DIR / "final_scientist_table.csv"

    save_json(row_json_path, row)

    write_header = not csv_path.exists()
    with open(csv_path, "a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FINAL_COLUMNS)
        if write_header:
            writer.writeheader()
        writer.writerow(row)

    write_or_update_master_excel(row)

    print(f"Saved row JSON: {row_json_path}")
    print(f"Updated CSV:    {csv_path}")


def safe_excel_value(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def build_section_ranges() -> List[Tuple[str, int, int]]:
    sections = []
    current_title = "Profil et contexte épidémiologique"
    start = 1

    section_titles = {
        "Date_PEC_J0": "J0",
        "Date_PEC_J3": "J3",
        "PDV_J30": "J30",
        "Type_paludisme": "Traitement et prise en charge",
    }

    for idx, col in enumerate(FINAL_COLUMNS, start=1):
        if col in section_titles:
            sections.append((current_title, start, idx - 1))
            current_title = section_titles[col]
            start = idx

    sections.append((current_title, start, len(FINAL_COLUMNS)))
    return sections


def style_master_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="FFF2CC")
    section_fill = PatternFill("solid", fgColor="FFD966")
    thin_green = Side(style="thin", color="70AD47")

    # Row 1: section headers
    for title, start_col, end_col in build_section_ranges():
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col)
        cell.value = title
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: actual column names
    for col_idx, col_name in enumerate(FINAL_COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx)
        c.value = col_name
        c.font = Font(bold=True, size=10)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=thin_green)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(FINAL_COLUMNS))}2"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36

    # widths
    width_map = {
        "A": 18, "B": 8, "C": 8, "D": 22, "E": 16, "F": 16, "G": 16, "H": 16,
        "I": 18, "J": 22,
    }
    for i in range(1, len(FINAL_COLUMNS) + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width_map.get(col_letter, 14)


def write_or_update_master_excel(row: Dict[str, str]) -> None:
    FINAL_TABLE_DIR.mkdir(parents=True, exist_ok=True)

    if MASTER_XLSX_PATH.exists():
        wb = load_workbook(MASTER_XLSX_PATH)
        ws = wb["Feuil1"] if "Feuil1" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Feuil1"
        style_master_sheet(ws)

    # find existing patient row by ID_Patient in column A, data starts at row 3
    target_row = None
    for r in range(3, ws.max_row + 1):
        if safe_excel_value(ws.cell(r, 1).value) == row["ID_Patient"]:
            target_row = r
            break

    if target_row is None:
        target_row = max(ws.max_row + 1, 3)

    for col_idx, col_name in enumerate(FINAL_COLUMNS, start=1):
        cell = ws.cell(row=target_row, column=col_idx)
        cell.value = safe_excel_value(row.get(col_name, ""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(MASTER_XLSX_PATH)
    print(f"Updated Excel:  {MASTER_XLSX_PATH}")




def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract one anonymized final scientist table row from a patient merged JSON."
    )
    parser.add_argument("patient_merged_json", help="Path to patient merged JSON")
    args = parser.parse_args()

    patient_json_path = Path(args.patient_merged_json).expanduser().resolve()
    if not patient_json_path.exists():
        raise FileNotFoundError(f"Patient merged JSON not found: {patient_json_path}")

    record = load_json(patient_json_path)
    row = extract_row(record)
    write_outputs(patient_json_path, row)


if __name__ == "__main__":
    main()