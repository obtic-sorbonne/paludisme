from pathlib import Path
import argparse
import re
import csv
import html
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


EXPECTED_HEADER = ["Description", "Résultat", "Unité", "Valeurs normales", "Val."]


def clean_text(s: str) -> str:
    s = str(s).strip()
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def norm(s: str) -> str:
    return clean_text(
        s.lower()
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ï", "i")
        .replace("î", "i")
        .replace("’", "'")
    )


def parse_final_txt(path: Path):
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    section = None
    table_rows = []
    notes = []

    for line in lines:
        line = line.rstrip()

        if line.startswith("=== FINAL TABLE"):
            section = "table"
            continue
        if line.startswith("=== FINAL NOTES"):
            section = "notes"
            continue
        if not line.strip():
            continue

        if section == "table" and "|" in line:
            parts = [clean_text(x) for x in line.split("|")]
            if len(parts) < 5:
                parts += [""] * (5 - len(parts))
            table_rows.append(parts[:5])

        elif section == "notes":
            notes.append(clean_text(line))

    return table_rows, notes


def ensure_header(rows):
    if not rows:
        return [EXPECTED_HEADER]

    first = rows[0]
    if len(first) < 5:
        first += [""] * (5 - len(first))
    first = first[:5]

    first_norm = [norm(x) for x in first]
    expected_norm = [norm(x) for x in EXPECTED_HEADER]

    if first_norm == expected_norm:
        rows[0] = EXPECTED_HEADER
        return rows

    return [EXPECTED_HEADER] + rows


def fix_table_rows(rows):
    fixed = []

    for i, row in enumerate(rows):
        desc, result, unit, normal, val = row

        if i == 0 and norm(desc) == "description":
            fixed.append(EXPECTED_HEADER)
            continue

        unit = unit.replace("g/di", "g/dl")
        unit = unit.replace("ux3", "μ×3")

        pct_names = {
            "hematocrite",
            "inddistriberythrocytair",
            "erythroblastes.",
            "polyneutrophiles.",
            "polyeosinophiles.",
            "polybasophiles.",
            "myelocytes.",
            "metamyelocytes",
            "lymphocytes.",
            "monocytes",
            "autres cellules.",
            "blastes",
            "plasmocytes",
        }

        if not unit and norm(desc) in pct_names:
            unit = "%"

        fixed.append([desc, result, unit, normal, val])

    return fixed


def split_note_sections(notes):
    section_titles = []
    cleaned_notes = []

    for line in notes:
        n = norm(line)
        if "examens d'hematologie" in n or n == "cytologie":
            section_titles.append(line)
        else:
            cleaned_notes.append(line)

    return section_titles, cleaned_notes


def merge_note_lines(notes):
    merged = []
    buffer = ""

    anchors = [
        "rech parasites sang",
        "ag plasmodium",
        "notion de voyage recent",
        "pays d'origine",
        "pays d' origine",
        "negative",
    ]

    def flush():
        nonlocal buffer
        if buffer.strip():
            merged.append(clean_text(buffer))
        buffer = ""

    for line in notes:
        t = clean_text(line)
        nt = norm(t)

        if any(a in nt for a in anchors):
            flush()
            merged.append(t)
            continue

        if not buffer:
            buffer = t
        else:
            buffer += " " + t

    flush()
    return merged


def ensure_rech_parasites(notes):
    joined = " ".join(norm(x) for x in notes)
    if "rech parasites sang" not in joined:
        notes.insert(0, "Rech parasites sang")
    return notes


def load_checkbox_json(path: Path | None):
    if path is None:
        return {"table_checked_rows": [], "choice_groups": {}}

    data = json.loads(path.read_text(encoding="utf-8"))

    table_checked_rows = data.get("table_checked_rows", [])
    choice_groups = data.get("choice_groups", {})

    table_checked_rows = [clean_text(x) for x in table_checked_rows if clean_text(x)]

    cleaned_groups = {}
    for k, vals in choice_groups.items():
        key = clean_text(k)
        if not key:
            continue
        if isinstance(vals, str):
            vals = [vals]
        vals = [clean_text(v) for v in vals if clean_text(v)]
        if vals:
            cleaned_groups[key] = vals

    return {
        "table_checked_rows": table_checked_rows,
        "choice_groups": cleaned_groups,
    }


def filter_table_rows_by_checkboxes(rows, checkbox_data):
    checked = checkbox_data.get("table_checked_rows", [])
    if not checked:
        return rows

    checked_norm = {norm(x) for x in checked}

    filtered = []
    for i, row in enumerate(rows):
        if i == 0:
            filtered.append(row)
            continue

        desc = row[0]
        if norm(desc) in checked_norm:
            filtered.append(row)

    return filtered


def append_choice_groups_to_notes(notes, checkbox_data):
    choice_groups = checkbox_data.get("choice_groups", {})
    if not choice_groups:
        return notes

    out = list(notes)

    if out and out[-1].strip():
        out.append("")

    out.append("=== SELECTED OPTIONS ===")

    for field, values in choice_groups.items():
        if not values:
            continue
        if len(values) == 1:
            out.append(f"{field} {values[0]}")
        else:
            out.append(f"{field} " + "; ".join(values))

    return out


def write_csv(rows, section_titles, notes, out_csv: Path, title: str):
    csv_rows = []

    csv_rows.append([title, "", "", "", ""])
    csv_rows.append(["", "", "", "", ""])

    for st in section_titles:
        csv_rows.append([st, "", "", "", ""])

    if section_titles:
        csv_rows.append(["", "", "", "", ""])

    csv_rows.extend(rows)

    if notes:
        csv_rows.append(["", "", "", "", ""])
        csv_rows.append(["Notes / Comment Section", "", "", "", ""])
        for note in notes:
            if note.strip():
                csv_rows.append([note, "", "", "", ""])

    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(csv_rows)

def write_xlsx(rows, section_titles, notes, out_xlsx: Path, title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    current_row = 1

    ws.cell(row=current_row, column=1, value=title)
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 2

    if section_titles:
        for st in section_titles:
            ws.cell(row=current_row, column=1, value=st)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
        current_row += 1

    table_start_row = current_row

    for r_idx, row in enumerate(rows, start=table_start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    header_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")

    for cell in ws[table_start_row]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(
        min_row=table_start_row + 1,
        max_row=table_start_row + len(rows) - 1,
        min_col=1,
        max_col=5,
    ):
        for i, cell in enumerate(row, start=1):
            if i == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "A": 42,
        "B": 14,
        "C": 16,
        "D": 18,
        "E": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    notes_start = table_start_row + len(rows) + 2
    ws.cell(row=notes_start, column=1, value="Notes / Comment Section")
    ws.cell(row=notes_start, column=1).font = Font(bold=True, size=12)

    note_row = notes_start + 1
    for note in notes:
        ws.cell(row=note_row, column=1, value=note)
        note_row += 1

    wb.save(out_xlsx)


def write_html(rows, section_titles, notes, out_html: Path, title: str):
    header_row = rows[0] if rows else EXPECTED_HEADER
    body_rows = rows[1:] if len(rows) > 1 else []

    header_html = "<tr>" + "".join(
        f"<th>{html.escape(cell)}</th>" for cell in header_row
    ) + "</tr>"

    body_html = []
    for row in body_rows:
        body_html.append(
            "<tr>" + "".join(f"<td>{html.escape(cell)}</td>" for cell in row) + "</tr>"
        )

    titles_html = ""
    if section_titles:
        titles_html = "<div class='section-titles'>" + "".join(
            f"<p><strong>{html.escape(t)}</strong></p>" for t in section_titles
        ) + "</div>"

    notes_html = ""
    if notes:
        notes_html = "<div class='notes'>" + "".join(
            f"<p>{html.escape(n)}</p>" for n in notes if n != ""
        ) + "</div>"

    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{html.escape(title)}</title>
<style>
body {{
    font-family: Arial, sans-serif;
    margin: 24px;
    color: #222;
    background: #ffffff;
}}

h1 {{
    font-size: 22px;
    margin-bottom: 18px;
}}

h2 {{
    margin-top: 28px;
    font-size: 18px;
}}

.table-wrap {{
    display: inline-block;
    border: 1px solid #bfbfbf;
    background: #fff;
}}

table {{
    border-collapse: collapse;
    table-layout: fixed;
    width: 100%;
    min-width: 900px;
}}

th, td {{
    border: 1px solid #c8c8c8;
    padding: 8px 10px;
    vertical-align: middle;
    word-wrap: break-word;
    text-align: center;
    font-size: 14px;
}}

th {{
    background: #efefef;
    font-weight: 600;
}}

td:first-child,
th:first-child {{
    text-align: left;
}}

colgroup col:nth-child(1) {{ width: 42%; }}
colgroup col:nth-child(2) {{ width: 14%; }}
colgroup col:nth-child(3) {{ width: 16%; }}
colgroup col:nth-child(4) {{ width: 18%; }}
colgroup col:nth-child(5) {{ width: 10%; }}

.section-titles {{
    margin: 18px 0 10px 0;
}}

.section-titles p {{
    margin: 0 0 6px 0;
}}

.notes {{
    border: 1px solid #c8c8c8;
    padding: 12px;
    background: #fafafa;
    max-width: 900px;
}}

.notes p {{
    margin: 0 0 10px 0;
}}
</style>
</head>
<body>
    <h1>{html.escape(title)}</h1>

    {titles_html}

    <div class="table-wrap">
        <table>
            <colgroup>
                <col><col><col><col><col>
            </colgroup>
            <thead>
                {header_html}
            </thead>
            <tbody>
                {''.join(body_html)}
            </tbody>
        </table>
    </div>

    <h2>Notes / Comment Section</h2>
    {notes_html}
</body>
</html>
"""
    out_html.write_text(page, encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Improve final hybrid table output")
    parser.add_argument("final_txt", help="Path to final hybrid cleaned txt")
    parser.add_argument(
        "--checkbox-json",
        default=None,
        help="Optional JSON file containing selected checkbox/radio results",
    )
    parser.add_argument(
        "--output-dir",
        default="/home/lfarooq/digitize_medical_records/benchmark_outputs/formatted_tables_improved",
        help="Directory to save improved outputs",
    )
    args = parser.parse_args()

    final_txt = Path(args.final_txt)
    checkbox_json = Path(args.checkbox_json) if args.checkbox_json else None

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    rows, notes = parse_final_txt(final_txt)
    checkbox_data = load_checkbox_json(checkbox_json)

    rows = ensure_header(rows)
    rows = fix_table_rows(rows)
    rows = filter_table_rows_by_checkboxes(rows, checkbox_data)

    section_titles, notes = split_note_sections(notes)
    notes = ensure_rech_parasites(notes)
    notes = merge_note_lines(notes)
    notes = append_choice_groups_to_notes(notes, checkbox_data)

    stem = final_txt.stem
    title = stem + "_improved"

    out_csv = out_dir / f"{stem}_improved.csv"
    out_html = out_dir / f"{stem}_improved.html"
    out_xlsx = out_dir / f"{stem}_improved.xlsx"

    write_csv(rows, section_titles, notes, out_csv, title)    
    write_html(rows, section_titles, notes, out_html, title)
    write_xlsx(rows, section_titles, notes, out_xlsx, title)

    print(f"Saved improved CSV:  {out_csv}")
    print(f"Saved improved HTML: {out_html}")
    print(f"Saved improved XLSX: {out_xlsx}")


if __name__ == "__main__":
    main()